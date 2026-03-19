#!/usr/bin/env python3
"""
CoFAST: COchlear Function Analysis SuiTe
=========================================
Developed by the Functional Testing Core at Eaton-Peabody Laboratories (EPL).
Integrates with the EPL CFTS acquisition system.

GUI application for analyzing auditory brainstem response (ABR) and
distortion product otoacoustic emission (DPOAE) data.

Features:
  1. Extract ABR thresholds → Excel sheet
  2. Extract Wave 1 growth functions → Excel sheets (one per frequency)
  3. Plot mean ± SEM waveform traces for selected mice/frequencies/levels
  4. Extract DPOAE thresholds using a user-specified SNR criterion → Excel sheet

Dependencies (install via pip):
  pip install numpy pandas openpyxl matplotlib scipy
"""

import os
import re
import sys
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('TkAgg')
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def _mk_btn(parent, text, command, bg, fg='white', **kw):
    """Cross-platform colored button.
    On macOS tk.Button ignores bg/fg; use a styled Label instead."""
    if sys.platform == 'darwin':
        btn = tk.Label(parent, text=text, bg=bg, fg=fg, cursor='hand2', **kw)
        btn.bind('<Button-1>', lambda e: command())
        btn.bind('<Enter>',   lambda e: btn.config(relief=tk.SUNKEN))
        btn.bind('<Leave>',   lambda e: btn.config(relief=tk.FLAT))
        btn.config(relief=tk.FLAT)
        return btn
    else:
        return tk.Button(parent, text=text, command=command,
                         bg=bg, fg=fg, relief=tk.FLAT, **kw)


# ============================================================
# Data Parsing Functions
# ============================================================

def parse_analyzed_file(filepath):
    """
    Parse an ABR *-analyzed.txt file produced by the EPL peak-analysis tool.

    Returns a dict:
        threshold   : float (dB SPL) or None
        frequency   : float (kHz) or None
        data        : pd.DataFrame with columns:
                        Level, 0.3msec Avg, 0.3msec StDev,
                        P1 Latency, P1 Amplitude, N1 Latency, N1 Amplitude,
                        … P5/N5 …, CorrCoef
        filepath    : original path
    """
    result = {'threshold': None, 'frequency': None, 'data': None, 'filepath': filepath}

    try:
        with open(filepath, 'r', errors='replace') as fh:
            lines = fh.readlines()
    except Exception:
        return result

    for line in lines[:6]:
        stripped = line.strip()
        if stripped.startswith('Threshold (dB SPL):'):
            try:
                result['threshold'] = float(stripped.split(':', 1)[1].strip())
            except ValueError:
                pass
        elif stripped.startswith('Frequency (kHz):'):
            try:
                result['frequency'] = float(stripped.split(':', 1)[1].strip())
            except ValueError:
                pass

    # Find the tab-delimited header row that starts with "Level"
    header_idx = None
    for i, line in enumerate(lines):
        if line.startswith('Level\t'):
            header_idx = i
            break
    if header_idx is None:
        return result

    headers = lines[header_idx].strip().split('\t')
    # Remove trailing empty column that the analysis tool sometimes adds
    while headers and headers[-1] == '':
        headers.pop()

    data_rows = []
    for line in lines[header_idx + 1:]:
        line = line.strip()
        if not line:
            continue
        parts = line.split('\t')
        try:
            row = [float(x) if x.strip() not in ('', 'nan', 'NaN') else np.nan
                   for x in parts]
            # Pad to match header length
            while len(row) < len(headers):
                row.append(np.nan)
            data_rows.append(row[:len(headers)])
        except ValueError:
            pass

    if data_rows:
        result['data'] = pd.DataFrame(data_rows, columns=headers)

    return result


def _abr_empty():
    return {'frequency': None, 'sample_rate_us': 40.0, 'levels': [],
            'waveforms': None, 'diff_waveforms': None}


def _parse_cfts_abr(content, filepath):
    """
    CFTS extensionless ABR / VsEP format.
    Data is stored as [sum | difference] — only the first half (sum) is the
    actual averaged waveform.  Levels can be negative (VsEP).
    """
    result = _abr_empty()
    if ':DATA' not in content:
        return result

    header_part, data_part = content.split(':DATA', 1)

    # Frequency (skip non-numeric stimulus descriptions like 'clicks', 'chirp', paths)
    fn = os.path.basename(filepath)
    if fn.lower().startswith('vsep'):
        result['frequency'] = None          # VsEP — no tone frequency
    else:
        m = re.search(r'SW FREQ:\s*([\d.]+)', header_part)
        if m:
            try:
                result['frequency'] = float(m.group(1))
            except ValueError:
                pass

    # Sample interval (µs/sample)
    m = re.search(r'SAMPLE\s*\(.*?sec\):\s*([\d.]+)', header_part)
    if m:
        result['sample_rate_us'] = float(m.group(1))

    # Levels — parse ALL entries including -Inf control markers.
    # We must reshape the data with the full column count before filtering.
    m = re.search(r':LEVELS:([\-\d;.\s Inf]+)', header_part)
    if not m:
        return result

    lev_str   = m.group(1).strip(';')
    all_levels = []
    for tok in lev_str.split(';'):
        tok = tok.strip()
        if not tok:
            continue
        try:
            all_levels.append(float(tok))   # handles '-Inf', 'Inf', '-15', etc.
        except ValueError:
            continue

    n_all = len(all_levels)
    if n_all == 0:
        return result

    # Parse flat number array using the FULL column count
    nums = np.array(re.findall(r'[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?', data_part),
                    dtype=np.float64)
    n_rows = len(nums) // n_all
    if n_rows == 0:
        return result

    mat = nums[:n_rows * n_all].reshape(n_rows, n_all)

    # CFTS files store sum+difference concatenated: take first half (the average)
    n_samples = n_rows // 2
    if n_samples == 0:
        n_samples = n_rows          # single-polarity file — use everything
        diff_half = None
    else:
        diff_half = mat[n_samples:, :]  # second half = polarity-difference channel
    mat = mat[:n_samples, :]

    # Drop -Inf / +Inf control columns (masker / noise-floor channels)
    keep = [i for i, lv in enumerate(all_levels) if np.isfinite(lv)]
    result['levels']         = [all_levels[i] for i in keep]
    result['waveforms']      = mat[:, keep]   # µV already
    result['diff_waveforms'] = diff_half[:, keep] if diff_half is not None else None
    return result


def _parse_fast_abr(content, filepath):
    """
    CFTS3 [FAST ABR] and [STANDARD ABR] .tsv / .txt formats.
    Data section begins after [DATA], first line is tab-separated column headers.
    Neural_* columns hold the averaged ABR; CM_* columns are cochlear microphonic.
    """
    result = _abr_empty()
    if '[DATA]' not in content:
        return result

    header_part, data_part = content.split('[DATA]', 1)

    # Frequency
    m = re.search(r'Frequency \(kHz\)=([\d.]+)', header_part)
    if not m:
        m = re.search(r'Stimulus\.Frequency \(kHz\)=([\d.]+)', header_part)
    if m:
        result['frequency'] = float(m.group(1))

    # Sampling rate → µs/sample
    m = re.search(r'Response\.Fs \(Hz\)=([\d.]+)', header_part)
    if not m:
        m = re.search(r'Response\.Sampling rate \(Hz\)=([\d.]+)', header_part)
    if m:
        result['sample_rate_us'] = 1e6 / float(m.group(1))

    # Levels
    m = re.search(r'Levels=([\-\d;. Inf]+)', header_part)
    if m:
        lev_str = m.group(1).strip(';')
        result['levels'] = [float(x) for x in lev_str.split(';')
                            if x.strip() and x.strip() not in ('Inf', '-Inf', '')]

    n_levels = len(result['levels'])
    if n_levels == 0:
        return result

    # Data lines: skip blank lines, first non-blank is column header
    lines = [l for l in data_part.split('\n') if l.strip()]
    if len(lines) < 2:
        return result

    col_names = [c.strip() for c in lines[0].split('\t')]
    neural_idx = [i for i, c in enumerate(col_names) if c.startswith('Neural_')]
    if not neural_idx:
        # Fall back: skip first (time) column, take next n_levels
        neural_idx = list(range(1, n_levels + 1))

    rows = []
    for line in lines[1:]:
        parts = line.split()
        if len(parts) <= max(neural_idx):
            continue
        try:
            rows.append([float(parts[i]) for i in neural_idx])
        except ValueError:
            continue

    if rows:
        result['waveforms'] = np.array(rows, dtype=np.float64)  # µV
    return result


def _parse_caspary_txt(content, filepath):
    """
    FBN / Caspary multi-level .txt format starting with 'Identifier:'.
    Intensity levels, sample period, stimulus frequency, and zero-position
    are all encoded as repeated comma-separated metadata rows.
    """
    result = _abr_empty()
    try:
        lev_str  = re.search(r'Intensity:([\d,]+)', content).group(1)
        levels   = [float(x) for x in re.findall(r',+([\d]+)', lev_str)]
        result['levels'] = levels

        dt_str   = re.search(r'Smp\. Period:([\d\.,]+)', content).group(1)
        dt_vals  = [float(x) for x in re.findall(r',+([\d.]+)', dt_str)]
        result['sample_rate_us'] = dt_vals[0]

        m_freq = re.search(r'Stim\. Freq[.,]*([\d,]+)', content)
        if m_freq:
            freqs = [float(x) for x in re.findall(r'[\d]+', m_freq.group(1))]
            if freqs:
                result['frequency'] = freqs[0] / 1000.0  # Hz → kHz

        zero_str = re.search(r'Zero Position:([\d,]+)', content).group(1)
        izero    = int(re.findall(r',+([\d]+)', zero_str)[0])

        # Data starts after 'Data Pnt' header line
        data_str = content.split('Data Pnt')[1].split('\n', 1)[1]
        a        = np.array(data_str.replace(',', ' ').split(), dtype=np.float64)
        n_cols   = len(levels) * 6 + 1
        y        = a.reshape(int(len(a) / n_cols), n_cols).T

        y = y[:, izero:]                  # start at zero-position offset
        data = y[2::6, :]                  # Average(uV) columns

        result['waveforms'] = data.T       # (n_samples, n_levels) µV
    except Exception:
        pass
    return result


def _parse_tabsep_txt(content, filepath):
    """
    Generic tab-separated .txt with a header row.
    First column = time (seconds).  Remaining columns are labelled with dB
    values in patterns like '8kHz80dB', '90 dBSPL', '80 dB', etc.
    """
    result = _abr_empty()
    lines = content.strip().split('\n')
    if len(lines) < 2:
        return result

    cols = [c.strip() for c in lines[0].split('\t')]
    levels = []
    col_idx = []
    for i, col in enumerate(cols[1:], 1):     # skip time column
        m = (re.search(r'(\d+)\s*dBSPL', col, re.I) or
             re.search(r'(\d+)\s*dB',    col, re.I) or
             re.search(r'kHz(\d+)dB',    col, re.I))
        if m:
            levels.append(float(m.group(1)))
            col_idx.append(i)

    if not levels:
        return result
    result['levels'] = levels

    time_vals, data_rows = [], []
    for line in lines[1:]:
        if not line.strip():
            continue
        parts = line.replace(',', ' ').split()
        if len(parts) <= max(col_idx):
            continue
        try:
            time_vals.append(float(parts[0]))
            data_rows.append([float(parts[i]) for i in col_idx])
        except (ValueError, IndexError):
            continue

    if not data_rows:
        return result

    data = np.array(data_rows, dtype=np.float64)
    t    = np.array(time_vals)

    # Auto-scale: convert Volts → µV if values are very small
    if np.max(np.abs(data)) < 0.01:
        data = data * 1e6

    if len(t) > 1:
        result['sample_rate_us'] = (t[1] - t[0]) * 1e6   # s → µs

    result['waveforms'] = data
    return result


def _parse_csv_abr(content, filepath):
    """
    Single- or few-level CSV: 'Time, C, R, AVG' or similar.
    Time is in seconds; signal columns are in Volts → converted to µV.
    """
    result = _abr_empty()
    try:
        header_line, data_str = content.split('\n', 1)
        cols     = [c.strip() for c in header_line.split(',')]
        n_cols   = len(cols)

        nums = np.array(data_str.replace(',', ' ').split(), dtype=np.float64)
        n_rows = len(nums) // n_cols
        if n_rows == 0:
            return result
        mat = nums[:n_rows * n_cols].reshape(n_rows, n_cols)

        t = mat[:, 0]                      # seconds

        # Choose signal columns (skip time; if >4 cols skip extra time cols too)
        if n_cols > 4:
            sig = mat[:, 3:5]
        else:
            sig = mat[:, 1:]

        sig = sig * 1e6                    # V → µV

        if len(t) > 1:
            result['sample_rate_us'] = (t[1] - t[0]) * 1e6

        result['levels'] = list(range(sig.shape[1]))   # nominal 0, 1, … labels
        result['waveforms'] = sig
    except Exception:
        pass
    return result


def parse_raw_abr_file(filepath):
    """
    Unified parser for multiple ABR/VsEP file formats.

    Returns a dict:
        frequency       : float (kHz) or None
        sample_rate_us  : float (µs per sample, default 40)
        levels          : list of floats (dB SPL)
        waveforms       : np.ndarray (n_samples × n_levels) in µV, or None

    Supported formats:
        • CFTS extensionless  — ':LEVELS:' / ':DATA' markers (ABR-*, VsEP-*)
        • CFTS3 FAST ABR      — '[FAST ABR]' / '[STANDARD ABR]' + '[DATA]'
        • Tab-separated .txt  — header row with dBSPL level labels
        • Caspary .txt        — starts with 'Identifier:'
        • CSV                 — Time, C, R, AVG columns
    """
    try:
        with open(filepath, 'r', encoding='latin-1', errors='replace') as fh:
            content = fh.read()
    except Exception:
        return _abr_empty()

    _, ext = os.path.splitext(filepath)
    ext = ext.lower()

    if ext == '.csv':
        return _parse_csv_abr(content, filepath)

    if ext in ('.tsv', '') and (content.startswith('[FAST ABR]') or
                                 content.startswith('[STANDARD ABR]')):
        return _parse_fast_abr(content, filepath)

    if ext == '.tsv':
        return _parse_fast_abr(content, filepath)

    if ext == '.txt':
        if content.startswith('Identifier:'):
            return _parse_caspary_txt(content, filepath)
        return _parse_tabsep_txt(content, filepath)

    # Default: extensionless CFTS / VsEP
    return _parse_cfts_abr(content, filepath)


def parse_dpoae_file(filepath):
    """
    Parse a DP-* DPOAE data file.

    Returns a pd.DataFrame with columns:
        Level, f1_Hz, f2_Hz, f1_dB, f2_dB, DP_dB, DP_Noise_dB
    """
    rows = []
    in_data = False

    try:
        with open(filepath, 'r', errors='replace') as fh:
            for line in fh:
                stripped = line.strip()
                if stripped == ':DATA':
                    in_data = True
                    continue
                if not in_data:
                    continue
                if not stripped or stripped.startswith(':'):
                    continue
                parts = stripped.split()
                if len(parts) < 7:
                    continue
                try:
                    def safe_float(s):
                        if s.lower() == 'nan':
                            return np.nan
                        return float(s)
                    rows.append({
                        'Level':       safe_float(parts[0]),
                        'f1_Hz':       safe_float(parts[1]),
                        'f2_Hz':       safe_float(parts[2]),
                        'f1_dB':       safe_float(parts[3]),
                        'f2_dB':       safe_float(parts[4]),
                        'DP_dB':       safe_float(parts[5]),
                        'DP_Noise_dB': safe_float(parts[6]),
                    })
                except (ValueError, IndexError):
                    pass
    except Exception:
        pass

    return pd.DataFrame(rows) if rows else pd.DataFrame(
        columns=['Level', 'f1_Hz', 'f2_Hz', 'f1_dB', 'f2_dB', 'DP_dB', 'DP_Noise_dB'])


def _extract_mouse_id(filename):
    """
    Extract a mouse ID from a CFTS-style filename.
    Examples:
        ABR-52-3            → 'ABR-52'
        ABR-136-1-analyzed.txt → 'ABR-136'
        VsEP-1018-13        → 'VsEP-1018'
        DP-52-3             → 'DP-52'
    Returns None if the pattern doesn't match.
    """
    fn = os.path.basename(filename).replace('-analyzed.txt', '')
    fn = os.path.splitext(fn)[0]   # strip any extension
    m = re.match(r'^([\w]+-\d+)-', fn)
    return m.group(1) if m else None


def find_mouse_data(root_dir):
    """
    Recursively scan root_dir for mouse data.

    Mouse identity is determined by the filename prefix when possible
    (e.g. ABR-52-3 and ABR-52-4 both belong to mouse 'ABR-52').
    For files that don't follow the CFTS naming convention the containing
    folder name is used as a fallback mouse ID.

    Sub-folders of an already-detected folder are excluded to avoid
    surfacing archive sub-directories.

    Returns:
        dict  {mouse_id: {
            'analyzed_files': [sorted paths],
            'raw_abr_files':  [sorted paths],
            'dp_files':       [sorted paths],
            'unanalyzed_raw': [sorted paths],
            'folder':         path,
            'group':          parent-folder name,
        }}
    """
    from collections import defaultdict

    def _empty_entry(dirpath):
        return {
            'analyzed_files': [],
            'raw_abr_files':  [],
            'dp_files':       [],
            'unanalyzed_raw': [],
            'folder':         dirpath,
            'group':          'All',
        }

    # ── First pass: walk and assign every file to a (folder, mouse_id) bucket ──
    # Use defaultdict keyed by (dirpath, mouse_id)
    buckets = defaultdict(lambda: None)

    def get_bucket(dirpath, mid):
        key = (dirpath, mid)
        if buckets[key] is None:
            buckets[key] = _empty_entry(dirpath)
        return buckets[key]

    visited_dirs = set()

    for dirpath, _dirnames, filenames in os.walk(root_dir):
        visited_dirs.add(dirpath)
        folder_id = os.path.basename(dirpath)

        for f in filenames:
            fp = os.path.join(dirpath, f)
            mid = folder_id   # animal name always comes from the folder

            if f.endswith('-analyzed.txt'):
                get_bucket(dirpath, mid)['analyzed_files'].append(fp)

            elif re.match(r'^(ABR|VsEP)-\d+-\d+$', f):
                get_bucket(dirpath, mid)['raw_abr_files'].append(fp)

            elif re.match(r'^DP-\d+-\d+$', f):
                get_bucket(dirpath, mid)['dp_files'].append(fp)

            elif f.endswith('.tsv') and not f.endswith('-analyzed.txt'):
                get_bucket(dirpath, mid)['raw_abr_files'].append(fp)

            elif (f.endswith('.txt') and not f.endswith('-analyzed.txt')
                  and re.match(r'^(ABR|VsEP)', f)):
                # Standard CFTS-named .txt ABR files
                get_bucket(dirpath, mid)['raw_abr_files'].append(fp)

            elif (f.endswith('.txt') and not f.endswith('-analyzed.txt')
                  and not re.match(r'^(ABR|VsEP|DP)', f)):
                # Non-standard .txt — only include if content matches a known ABR signature
                try:
                    with open(fp, 'r', encoding='latin-1', errors='replace') as _fh:
                        _snippet = _fh.read(100)
                    if (_snippet.startswith('Identifier:') or
                            _snippet.startswith('[FAST ABR]') or
                            _snippet.startswith('[STANDARD ABR]') or
                            _snippet.startswith('time\t')):   # tab-separated waveform file
                        get_bucket(dirpath, mid)['raw_abr_files'].append(fp)
                except (OSError, IOError):
                    pass

            elif f.endswith('.csv'):
                get_bucket(dirpath, mid)['raw_abr_files'].append(fp)

    # ── Second pass: compute unanalyzed, sort lists ──
    # Exclude buckets from subdirectories of other populated directories
    all_dirs = sorted(visited_dirs)
    populated_dirs = {dirpath for (dirpath, _) in buckets if buckets[(dirpath, _)] is not None
                      and (buckets[(dirpath, _)]['analyzed_files'] or
                           buckets[(dirpath, _)]['raw_abr_files'] or
                           buckets[(dirpath, _)]['dp_files'])}

    excluded_dirs = set()
    for path in list(populated_dirs):
        for other in list(populated_dirs):
            if path != other and path.startswith(other + os.sep):
                excluded_dirs.add(path)

    mice = {}
    for (dirpath, mid), info in buckets.items():
        if info is None:
            continue
        if dirpath in excluded_dirs:
            continue
        if not (info['analyzed_files'] or info['raw_abr_files'] or info['dp_files']):
            continue

        # Compute unanalyzed: raw files with no corresponding -analyzed.txt
        analyzed_bases = {fp.replace('-analyzed.txt', '') for fp in info['analyzed_files']}
        info['unanalyzed_raw'] = [fp for fp in info['raw_abr_files']
                                   if fp not in analyzed_bases]
        info['raw_abr_files']  = sorted(info['raw_abr_files'])
        info['analyzed_files'] = sorted(info['analyzed_files'])
        info['dp_files']       = sorted(info['dp_files'])

        # Handle duplicate mouse IDs across folders: append folder suffix
        if mid in mice and mice[mid]['folder'] != dirpath:
            suffix = os.path.basename(dirpath)
            mid = f"{mid} ({suffix})"

        mice[mid] = info

    return mice


# ============================================================
# Signal processing helpers (used by peak-analysis window)
# ============================================================

def _butter_bandpass(signal_uv, fs_hz, fl=200.0, fh=10000.0):
    """
    Zero-phase 1st-order Butterworth bandpass filter.
    Returns (filtered_signal, zpk_description_string).
    Falls back to unfiltered signal if scipy is unavailable.
    """
    try:
        from scipy.signal import butter, filtfilt
        nyq  = 0.5 * fs_hz
        low  = fl  / nyq
        high = min(fh / nyq, 0.9999)
        z, p, k = butter(1, [low, high], btype='bandpass', output='zpk')
        b,  a   = butter(1, [low, high], btype='bandpass', output='ba')
        filtered  = filtfilt(b, a, signal_uv)
        zpk_str   = (f"Pass 0 -- z: {repr(z)}, p: {repr(p)}, "
                     f"k: {repr(np.float64(k))}")
        return filtered, zpk_str
    except Exception:
        return signal_uv.copy(), "No filtering"


def _snap_to_extremum(signal_uv, time_ms, t_click, find_max, half_win=0.5):
    """
    Return the index of the nearest local maximum (find_max=True) or
    minimum (find_max=False) within ±half_win ms of t_click.
    """
    i0 = max(0,               np.searchsorted(time_ms, t_click - half_win))
    i1 = min(len(signal_uv),  np.searchsorted(time_ms, t_click + half_win))
    if i0 >= i1:
        return int(np.clip(np.searchsorted(time_ms, t_click), 0,
                           len(signal_uv) - 1))
    seg = signal_uv[i0:i1]
    try:
        from scipy.signal import find_peaks as _sp_fp
        pks, props = _sp_fp(seg if find_max else -seg, prominence=0)
        if len(pks):
            return int(i0 + pks[np.argmax(props['prominences'])])
    except Exception:
        pass
    return int(i0 + (np.argmax(seg) if find_max else np.argmin(seg)))


def _cluster_by_amplitude(vals, spacing):
    """
    Group consecutive elements where |vals[i-1] - vals[i]| <= spacing.
    Returns list of lists of integer indices into vals.
    Port of peakdetect.py cluster_indices().
    """
    indices = [0]
    clusters = []
    for i in range(1, len(vals)):
        if abs(vals[i - 1] - vals[i]) <= spacing:
            indices.append(i)
        else:
            clusters.append(indices)
            indices = [i]
    clusters.append(indices)
    return clusters


def _cluster_by_position(positions, amplitudes, spacing):
    """
    Cluster peaks by spatial proximity; return index of max-amplitude peak
    in each cluster.  Port of peakdetect.py cluster().
    positions  : 1-D array of sample indices (sorted)
    amplitudes : waveform values at those positions
    spacing    : max sample distance to merge into same cluster
    Returns list of integer indices into the positions/amplitudes arrays.
    """
    clusters = _cluster_by_amplitude(positions, spacing)
    result = []
    for c in clusters:
        amps_c = amplitudes[c]
        best = c[int(np.where(amps_c == amps_c.max())[0][0])]
        result.append(best)
    return result


def _nzc_peaks(signal_uv, time_ms, min_latency_ms=1.0, min_spacing_ms=0.3, dev=1.0):
    """
    Faithful port of peakdetect.py nzc_noise_filtered + np_basic.

    Algorithm:
      1. Find all local maxima (p_ind) and local minima (n_ind).
      2. Combine and sort temporally → combined.
      3. cluster_by_amplitude(waveform[combined], min_noise): groups consecutive
         extrema whose amplitude values differ by <= min_noise.
      4. Pop the first cluster (baseline noise near zero).
      5. For each remaining cluster: if #peaks > #valleys → keep max peak.
      6. If >5 peaks remain, spatially deduplicate with min_spacing.
      7. Filter to after min_latency_ms.
    Returns list of sample indices in temporal order.
    """
    x = signal_uv
    dx = np.diff(x)
    p_ind = np.where((dx[1:] < 0) & (dx[:-1] >= 0))[0] + 1   # local maxima
    n_ind = np.where((dx[1:] > 0) & (dx[:-1] <= 0))[0] + 1   # local minima

    if len(p_ind) == 0:
        return []

    dt_ms = time_ms[1] - time_ms[0]
    fs_hz = 1000.0 / dt_ms
    baseline_samples = max(1, int(1e-3 * fs_hz))
    min_noise = x[:baseline_samples].std() * dev
    min_noise = max(min_noise, 1e-9)

    # Combine peaks + valleys sorted by position
    combined = np.sort(np.r_[p_ind, n_ind])

    # Cluster by amplitude proximity of consecutive extrema values
    clusters = _cluster_by_amplitude(x[combined], min_noise)

    # Discard first cluster (baseline / near-zero noise)
    if clusters:
        clusters.pop(0)

    p_ind_set = set(p_ind.tolist())
    n_ind_set = set(n_ind.tolist())

    ind = []
    for c in clusters:
        c_positions = combined[c]
        c_peaks   = [i for i in c_positions if i in p_ind_set]
        c_valleys = [i for i in c_positions if i in n_ind_set]
        if len(c_peaks) > len(c_valleys):
            best = int(np.where(x == x[c_peaks].max())[0][0])
            ind.append(best)

    ind = np.array(ind)

    # If >5, spatially deduplicate
    if len(ind) > 5:
        spacing_samples = min_spacing_ms * 1e-3 * fs_hz
        keep = _cluster_by_position(ind, x[ind], spacing_samples)
        ind = ind[keep]

    # Filter to after min_latency
    lb = min_latency_ms * 1e-3 * fs_hz
    ind = ind[ind >= lb]

    return ind.tolist()


def _auto_detect_peaks(signal_uv, time_ms, seed_peaks=None, seed_valleys=None):
    """
    Detect P1–P5 and N1–N5 in one ABR waveform.

    Unseeded: uses NZC-based algorithm (replicating peakdetect.py
    nzc_noise_filtered + np_basic) — takes n-th local maximum in
    temporal order after 1 ms.

    Seeded: narrows a ±SEED_HW ms window around the seed position
    and snaps to the local maximum within that window.

    Returns (peak_idx, valley_idx) – both dicts mapping {1-5 : int}.
    """
    SEED_HW = 0.7   # ±ms half-window when seeded

    # --- NZC-based significant peaks (unseeded path) ---
    nzc_peaks = _nzc_peaks(signal_uv, time_ms, min_latency_ms=1.0)

    peak_idx = {}
    for n in range(1, 6):
        wave_n = n - 1   # 0-indexed
        if seed_peaks and n in seed_peaks:
            # Seeded: snap to local max within ±SEED_HW of seed
            ct = time_ms[min(seed_peaks[n], len(time_ms) - 1)]
            peak_idx[n] = _snap_to_extremum(signal_uv, time_ms,
                                            ct, find_max=True,
                                            half_win=SEED_HW)
        elif wave_n < len(nzc_peaks):
            # Use n-th NZC peak in temporal order
            peak_idx[n] = int(nzc_peaks[wave_n])
        else:
            # Fallback: snap to max in a broad default window
            PEAK_WIN = {1: (1.0, 2.6), 2: (2.0, 3.6), 3: (2.8, 4.8),
                        4: (3.5, 5.8), 5: (4.5, 7.5)}
            t0, t1 = PEAK_WIN[n]
            peak_idx[n] = _snap_to_extremum(signal_uv, time_ms,
                                            (t0 + t1) / 2, find_max=True,
                                            half_win=(t1 - t0) / 2)

    valley_idx = {}
    for n in range(1, 6):
        pt   = time_ms[peak_idx[n]]
        v_lo = pt + 0.15
        v_hi = (time_ms[peak_idx[n + 1]] + 0.3) if n < 5 else (pt + 2.0)
        if seed_valleys and n in seed_valleys:
            ct   = time_ms[min(seed_valleys[n], len(time_ms) - 1)]
            v_lo = max(v_lo, ct - SEED_HW)
            v_hi = min(v_hi, ct + SEED_HW)
        valley_idx[n] = _snap_to_extremum(signal_uv, time_ms,
                                          (v_lo + v_hi) / 2, find_max=False,
                                          half_win=(v_hi - v_lo) / 2)
    return peak_idx, valley_idx


def _compute_corrcoef(y1, y2, time_ms, tmax=None):
    """Pearson r between two waveforms over 0 – tmax ms (full waveform if tmax is None)."""
    mask = time_ms <= (tmax if tmax is not None else time_ms[-1])
    a, b = y1[mask], y2[mask]
    if len(a) < 2 or np.std(a) == 0 or np.std(b) == 0:
        return np.nan
    return float(np.corrcoef(a, b)[0, 1])


_BASELINE_MS = 0.3   # ms window used for baseline statistics


def _write_analyzed_file(out_path, frequency_khz, threshold_db,
                         time_ms, waveform_mat, level_list,
                         peak_idx_list, valley_idx_list,
                         filter_zpk_str="No filtering"):
    """
    Write a *-analyzed.txt file in the EPL CFTS format.

    waveform_mat     : ndarray (n_samples × n_levels), µV,
                       columns ordered HIGHEST level first
    level_list       : list of floats, same column order
    peak_idx_list /
    valley_idx_list  : list of {1-5: int} dicts, same order
    """
    thr  = threshold_db   # may be None — written as NaN if so
    dt   = time_ms[1] - time_ms[0]
    n_bl = max(1, int(_BASELINE_MS / dt))

    # corrcoef[i] = r(level[i],  level[i-1]);  corrcoef[0] = NaN (no level above)
    corrcoefs = [np.nan]
    for i in range(1, len(level_list)):
        corrcoefs.append(
            _compute_corrcoef(waveform_mat[:, i], waveform_mat[:, i - 1],
                              time_ms))

    cols = ["Level", "0.3msec Avg", "0.3msec StDev"]
    for n in range(1, 6):
        cols += [f"P{n} Latency", f"P{n} Amplitude",
                 f"N{n} Latency", f"N{n} Amplitude"]
    cols.append("CorrCoef")

    thr_str = f"{thr:.0f}" if thr is not None else "NaN"
    with open(out_path, 'w') as fh:
        fh.write(f"Threshold (dB SPL): {thr_str}\n")
        fh.write(f"Frequency (kHz): {frequency_khz:.2f}\n")
        fh.write("Threshold estimation: manual\n")
        fh.write("Filter history (zpk format):\n")
        fh.write(filter_zpk_str + "\n")
        fh.write("NOTE: Negative latencies indicate no peak\n")
        fh.write("\t".join(cols) + "\t\n")

        for i, (lv, pidx, nidx) in enumerate(
                zip(level_list, peak_idx_list, valley_idx_list)):

            supra = (thr is None or lv >= thr)
            sign  = 1 if supra else -1
            bl    = waveform_mat[:n_bl, i]
            row   = [f"{lv:.2f}",
                     f"{float(np.mean(bl)):.6f}",
                     f"{float(np.std(bl)):.6f}"]

            for n in range(1, 6):
                pi = int(np.clip(pidx.get(n, 0), 0, len(time_ms) - 1))
                ni = int(np.clip(nidx.get(n, 0), 0, len(time_ms) - 1))
                row += [f"{sign * time_ms[pi]:.2f}",
                        f"{float(waveform_mat[pi, i]):.2f}",
                        f"{sign * time_ms[ni]:.2f}",
                        f"{float(waveform_mat[ni, i]):.2f}"]

            cc = corrcoefs[i]
            if i > 0 and not np.isnan(cc):
                row.append(f"{cc:.3f}")

            fh.write("\t".join(row) + "\n")


def dpoae_threshold(dp_df, f2_hz, criterion_db):
    """
    Return the DPOAE threshold (dB SPL) for a given f2 frequency.

    Threshold = lowest stimulus level where BOTH conditions are met:
      1. DP_dB >= criterion_db  (absolute amplitude)
      2. DP_dB - DP_Noise_dB >= 5.0 dB  (at least 5 dB above noise floor)
    Returns np.nan if no level meets both criteria (unmeasurable / NR).
    """
    mask = np.abs(dp_df['f2_Hz'] - f2_hz) < 200
    freq_data = dp_df[mask].copy().sort_values('Level')
    for _, row in freq_data.iterrows():
        if np.isnan(row['DP_dB']) or np.isnan(row['DP_Noise_dB']):
            continue
        if (row['DP_dB'] >= criterion_db and
                row['DP_dB'] - row['DP_Noise_dB'] >= 5.0):
            return row['Level']
    return np.nan


# ============================================================
# Main Application
# ============================================================

HEADER_BG   = '#1e4d7b'
BTN_BLUE    = '#2980b9'
BTN_GREEN   = '#27ae60'
BTN_PURPLE  = '#8e44ad'
BTN_ORANGE  = '#d35400'


def _pick_folders(parent=None, title="Select Folders"):
    """
    Open a native Windows multi-folder picker (IFileOpenDialog).
    The user can Ctrl+click or Shift+click to select multiple folders at once.
    Falls back to single-folder askdirectory on non-Windows or COM failure.
    Returns a list of selected folder paths (may be empty if cancelled).
    """
    import sys
    if sys.platform != 'win32':
        p = filedialog.askdirectory(title=title)
        return [p] if p else []

    import ctypes

    class GUID(ctypes.Structure):
        _fields_ = [('Data1', ctypes.c_uint32), ('Data2', ctypes.c_uint16),
                    ('Data3', ctypes.c_uint16), ('Data4', ctypes.c_uint8 * 8)]

    def make_guid(s):
        s = s.strip('{}').replace('-', '')
        g = GUID()
        g.Data1, g.Data2, g.Data3 = int(s[:8], 16), int(s[8:12], 16), int(s[12:16], 16)
        for i, b in enumerate(bytes.fromhex(s[16:])):
            g.Data4[i] = b
        return g

    try:
        ole32 = ctypes.windll.ole32
        ole32.CoInitialize(None)

        CLSID = make_guid('DC1C5A9C-E88A-4dde-A5A1-60F82A20AEF7')
        IID   = make_guid('D57C7288-D4AD-4768-BE02-9D969532D960')
        HR, DW = ctypes.HRESULT, ctypes.c_uint32

        dlg = ctypes.c_void_p()
        if ole32.CoCreateInstance(ctypes.byref(CLSID), None, 1,
                                  ctypes.byref(IID), ctypes.byref(dlg)) != 0:
            raise RuntimeError

        vt = ctypes.cast(ctypes.cast(dlg, ctypes.POINTER(ctypes.c_void_p))[0],
                         ctypes.POINTER(ctypes.c_void_p))

        GetOptions = ctypes.WINFUNCTYPE(HR, ctypes.c_void_p, ctypes.POINTER(DW))(vt[10])
        SetOptions = ctypes.WINFUNCTYPE(HR, ctypes.c_void_p, DW)(vt[9])
        SetTitle   = ctypes.WINFUNCTYPE(HR, ctypes.c_void_p, ctypes.c_wchar_p)(vt[17])
        Show       = ctypes.WINFUNCTYPE(HR, ctypes.c_void_p, ctypes.c_void_p)(vt[3])
        GetResults = ctypes.WINFUNCTYPE(HR, ctypes.c_void_p,
                                        ctypes.POINTER(ctypes.c_void_p))(vt[27])
        Release    = ctypes.WINFUNCTYPE(HR, ctypes.c_void_p)(vt[2])

        opts = DW(0)
        GetOptions(dlg, ctypes.byref(opts))
        SetOptions(dlg, opts.value | 0x20 | 0x200)   # FOS_PICKFOLDERS | FOS_ALLOWMULTISELECT
        SetTitle(dlg, title)

        hwnd = ctypes.c_void_p(parent.winfo_id() if parent else 0)
        hr = Show(dlg, hwnd)

        paths = []
        if hr == 0:
            arr = ctypes.c_void_p()
            if GetResults(dlg, ctypes.byref(arr)) == 0:
                vt_a = ctypes.cast(ctypes.cast(arr, ctypes.POINTER(ctypes.c_void_p))[0],
                                   ctypes.POINTER(ctypes.c_void_p))
                GetCount    = ctypes.WINFUNCTYPE(HR, ctypes.c_void_p, ctypes.POINTER(DW))(vt_a[7])
                GetItemAt   = ctypes.WINFUNCTYPE(HR, ctypes.c_void_p, DW,
                                                 ctypes.POINTER(ctypes.c_void_p))(vt_a[8])
                Release_arr = ctypes.WINFUNCTYPE(HR, ctypes.c_void_p)(vt_a[2])

                n = DW(0)
                GetCount(arr, ctypes.byref(n))
                for i in range(n.value):
                    item = ctypes.c_void_p()
                    if GetItemAt(arr, DW(i), ctypes.byref(item)) == 0:
                        vt_i = ctypes.cast(ctypes.cast(item, ctypes.POINTER(ctypes.c_void_p))[0],
                                           ctypes.POINTER(ctypes.c_void_p))
                        GetDisplayName = ctypes.WINFUNCTYPE(
                            HR, ctypes.c_void_p, ctypes.c_int,
                            ctypes.POINTER(ctypes.c_void_p))(vt_i[5])
                        Release_item = ctypes.WINFUNCTYPE(HR, ctypes.c_void_p)(vt_i[2])

                        buf = ctypes.c_void_p()
                        # SIGDN_FILESYSPATH = 0x80058000 (-2147319808 signed)
                        if GetDisplayName(item, ctypes.c_int(-2147319808),
                                          ctypes.byref(buf)) == 0 and buf.value:
                            paths.append(ctypes.wstring_at(buf.value))
                            ole32.CoTaskMemFree(buf)
                        Release_item(item)
                Release_arr(arr)

        Release(dlg)
        return paths

    except Exception:
        p = filedialog.askdirectory(title=title)
        return [p] if p else []


class ABRAnalysisApp:
    def __init__(self, root):
        self.root = root
        self.root.title("CoFAST: COchlear Function Analysis SuiTe")
        self.root.geometry("1150x780")
        self.root.minsize(900, 600)

        # Data storage
        self.mice_data       = {}   # {mouse_id: {...}}
        self.analyzed_cache  = {}   # {filepath: parsed dict}
        self._thresh_raw_map = {}   # {(mid, freq_khz): raw_path}
        self._wave1_raw_map  = {}   # {mid: raw_path} for current wave1 freq
        self.raw_cache       = {}   # {filepath: parsed dict}
        self.dp_cache        = {}   # {filepath: parsed dict}
        self._group_overrides = {}  # {mouse_id: str} — user-edited group labels

        self._build_ui()

    # ----------------------------------------------------------
    # UI Construction
    # ----------------------------------------------------------

    def _build_ui(self):
        # ── Menu bar ──────────────────────────────────────────
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(label="Add Data…", command=self._add_data)
        file_menu.add_command(label="Clear All Data",               command=self.clear_all_data)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.root.quit)
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Help", menu=help_menu)
        help_menu.add_command(label="About", command=self._show_about)

        # ── Top banner ────────────────────────────────────────
        banner = tk.Frame(self.root, bg=HEADER_BG, pady=7)
        banner.pack(fill=tk.X)
        tk.Label(banner, text="CoFAST: COchlear Function Analysis SuiTe",
                 font=('Arial', 13, 'bold'), bg=HEADER_BG, fg='white').pack(side=tk.LEFT, padx=12)
        self.folder_label = tk.Label(banner, text="No data loaded",
                                     font=('Arial', 9), bg=HEADER_BG, fg='#aad4f5')
        self.folder_label.pack(side=tk.LEFT, padx=6)
        _mk_btn(banner, text="🗑  Clear All", command=self.clear_all_data,
                bg='#c0392b', padx=10, pady=2).pack(side=tk.RIGHT, padx=(0, 6))
        _mk_btn(banner, text="✖  Remove Session", command=self.remove_animals,
                bg='#e67e22', padx=10, pady=2).pack(side=tk.RIGHT, padx=(0, 4))
        _mk_btn(banner, text="➕  Add Data", command=self._add_data,
                bg='#27ae60', padx=10, pady=2).pack(side=tk.RIGHT, padx=10)

        # ── Notebook ──────────────────────────────────────────
        style = ttk.Style()
        style.configure('TNotebook.Tab', padding=(12, 4), font=('Arial', 9))

        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=6, pady=4)

        self._build_overview_tab()
        self._build_threshold_tab()
        self._build_wave1_tab()
        self._build_latency_tab()
        self._build_plot_tab()
        self._build_dpoae_tab()

        # Auto-refresh whichever data tab becomes visible
        self._tab_names = ['Overview', 'ABR Thresholds', 'Wave Growth',
                           'Latencies', 'Plot Traces', 'DPOAE Thresholds']
        self._tab_refresh = {
            'ABR Thresholds':    self.refresh_threshold_table,
            'Wave Growth':       self.refresh_wave1_table,
            'Latencies':         self.refresh_latency_table,
            'DPOAE Thresholds':  self.refresh_dpoae_table,
        }
        self.notebook.bind('<<NotebookTabChanged>>', self._on_tab_change)

    # ── Tab 0: Overview ───────────────────────────────────────

    def _build_overview_tab(self):
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="  Overview  ")

        # Single scrollable Text widget for the whole tab (guide + live summary)
        outer = tk.Frame(frame)
        outer.pack(fill=tk.BOTH, expand=True, padx=10, pady=8)

        self.overview_text = tk.Text(
            outer, font=('Arial', 10), wrap=tk.WORD,
            state=tk.DISABLED, bg='#f9fafb', fg='#1a1a2e',
            padx=18, pady=10, relief=tk.FLAT,
            selectbackground='#cce5ff')
        vsb = ttk.Scrollbar(outer, command=self.overview_text.yview)
        self.overview_text.configure(yscrollcommand=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self.overview_text.pack(fill=tk.BOTH, expand=True)

        # ── Text tags ──────────────────────────────────────────────────
        T = self.overview_text
        T.tag_configure('h1',  font=('Arial', 14, 'bold'), foreground='#1a3a6b',
                        spacing1=14, spacing3=4)
        T.tag_configure('h2',  font=('Arial', 11, 'bold'), foreground='#1a3a6b',
                        spacing1=10, spacing3=2)
        T.tag_configure('body', font=('Arial', 10), foreground='#1a1a2e', spacing3=3)
        T.tag_configure('note', font=('Arial', 10, 'italic'), foreground='#555555',
                        spacing3=3)
        T.tag_configure('key',  font=('Courier', 10, 'bold'), foreground='#0d47a1',
                        background='#e8f0fe')
        T.tag_configure('bullet', font=('Arial', 10), foreground='#1a1a2e',
                        lmargin1=24, lmargin2=36, spacing3=2)
        T.tag_configure('code',   font=('Courier', 9),  foreground='#333333',
                        background='#eeeeee', lmargin1=32, lmargin2=32, spacing3=2)
        T.tag_configure('sep',    font=('Arial', 2),   spacing1=4, spacing3=4)
        T.tag_configure('sumhdr', font=('Arial', 10, 'bold'), foreground='#ffffff',
                        background='#2c3e50', spacing1=6, spacing3=6)
        T.tag_configure('sumtxt', font=('Courier', 9), foreground='#d4d4d4',
                        background='#1e1e1e', spacing3=1)

        # ── Helper to insert styled content ───────────────────────────
        def h1(txt):
            T.insert(tk.END, txt + '\n', 'h1')
        def h2(txt):
            T.insert(tk.END, txt + '\n', 'h2')
        def body(txt):
            T.insert(tk.END, txt + '\n', 'body')
        def note(txt):
            T.insert(tk.END, txt + '\n', 'note')
        def bullet(txt):
            T.insert(tk.END, '  •  ' + txt + '\n', 'bullet')
        def code(txt):
            T.insert(tk.END, txt + '\n', 'code')
        def blank():
            T.insert(tk.END, '\n', 'body')

        # ── Write the guide ────────────────────────────────────────────
        T.config(state=tk.NORMAL)
        T.delete(1.0, tk.END)

        h1("CoFAST: COchlear Function Analysis SuiTe  —  User Guide")
        note("Developed by the Functional Testing Core at Eaton-Peabody Laboratories (EPL). "
             "Integrates with the EPL CFTS acquisition system.")
        blank()

        # ── 1. LOADING DATA ───────────────────────────────────────────
        h2("1.  Loading Data")
        body("Click ➕ Add Data in the toolbar, then choose one of:")
        bullet("Add Folder(s) — select one or more animal data folders. "
               "The tool scans recursively and loads every ABR and DPOAE file it finds.")
        bullet("Open Files — select individual raw or analyzed ABR files directly.")
        body("You can load multiple folders in sequence; new animals are merged into "
             "the current dataset. Use ✖ Remove Session to drop a specific session, "
             "or 🗑 Clear All to start fresh.")
        blank()

        # ── 2. FOLDER STRUCTURE ───────────────────────────────────────
        h2("2.  Recommended Folder Structure")
        body("The tool derives animal IDs and group labels automatically from the "
             "folder hierarchy:")
        code("  Experiment root/")
        code("    WT Data/               ← group name comes from this folder")
        code("      Mouse001/            ← animal ID = this folder name")
        code("        ABR-4000-70        ← raw EPL CFTS waveform file")
        code("        ABR-4000-70-analyzed.txt  ← EPL peak-analysis output")
        code("        DP-1               ← DPOAE data file")
        code("      Mouse002/ …")
        code("    KO Data/")
        code("      Mouse003/ …")
        blank()
        body("Recognized file types:")
        bullet("EPL CFTS extensionless files  (ABR-####-##, VsEP-####-##)")
        bullet("EPL analyzed output files     (*-analyzed.txt)")
        bullet("Tab-separated waveform files  (*.tsv, or .txt files starting with 'time\\t')")
        bullet("CSV waveform files            (*.csv)")
        bullet("DPOAE data files              (DP-##)")
        blank()

        # ── 3. ANALYZED VS UNANALYZED ─────────────────────────────────
        h2("3.  Analyzed vs. Unanalyzed Files")
        body("The tool distinguishes two states for every recording:")
        bullet("Analyzed  (*-analyzed.txt present)  —  threshold, P1–P5 latencies, "
               "and N1–N5 latencies/amplitudes are already saved. These populate "
               "the ABR Thresholds, Wave Growth, and Latencies tabs immediately.")
        bullet("Unanalyzed  (raw file only, no matching -analyzed.txt)  —  "
               "double-clicking the file in the ABR Thresholds or Wave Growth table "
               "opens the Peak Analysis window so you can assign peaks and save. "
               "Once saved, the table refreshes automatically.")
        note("If you load an *-analyzed.txt file via Open Files, the tool will tell "
             "you to use Add Folder instead — analyzed files must be loaded as part "
             "of their containing folder so the animal ID is derived correctly.")
        blank()

        # ── 4. GROUPS ─────────────────────────────────────────────────
        h2("4.  Groups")
        body("Every session starts in the 'All' group so it is included in plots immediately. "
             "To reassign, double-click the Group cell in any table and type a new name "
             "(e.g. 'WT', 'KO', 'Noise-Exposed'). The change applies across all tabs.")
        body("Sessions with an empty group label or the label 'N/A' are excluded from "
             "all plots but still appear in the tables. Set Group to N/A (or clear it) "
             "to exclude a session from plots without removing it from the dataset.")
        body("Sessions that share a group name are averaged together in Mean ± SEM plots. "
             "Use the Individual plot mode to see every session overlaid.")
        blank()

        # ── 5. TABS ───────────────────────────────────────────────────
        h2("5.  Tabs at a Glance")
        bullet("ABR Thresholds  —  Threshold (dB SPL) × frequency table + audiogram plot. "
               "NM = not measured (no file for that frequency).")
        bullet("Wave Growth  —  Peak-to-trough amplitude (Pn − Nn, µV) × level table "
               "for any selected wave (1–5) and frequency. Plot shows amplitude growth functions.")
        bullet("Latencies  —  Peak latency (ms) × level table for any wave (1–5), "
               "P or N, and frequency. Plot shows latency-intensity functions.")
        bullet("Plot Traces  —  Overlay mean ± SEM waveforms for any combination of "
               "animals, frequency, and levels. Four layout modes: Stacked, Overlay, "
               "By Group, Mean + Individuals.")
        bullet("DPOAE Thresholds  —  DP threshold table + audiogram. "
               "Threshold = lowest level where DP ≥ criterion (dB SPL) "
               "AND DP is ≥ 5 dB above the noise floor.")
        note("All frequency axes use a log₂ scale (octave-linear: 4, 8, 16, 32 kHz …).")
        blank()

        # ── 6. PEAK ANALYSIS WINDOW ───────────────────────────────────
        h2("6.  Peak Analysis Window  —  Keyboard Reference")
        body("The peak analysis window opens when you double-click an unanalyzed row, "
             "or use Open Files on a raw ABR file. "
             "The current waveform is drawn as a thick black line. "
             "Waveforms at or above threshold are drawn in dark grey; "
             "sub-threshold waveforms are drawn as lighter dashed lines. "
             "The threshold level is marked with a horizontal dotted blue line. "
             "Peaks are colour-coded: red = P1/N1, orange = P2/N2, "
             "green = P3/N3, teal = P4/N4, blue = P5/N5. "
             "Positive peaks (P) are shown as circles; "
             "negative troughs (N) as inverted triangles.")
        blank()
        h2("   Navigating waveforms")
        bullet("↑ / ↓           Move to the next or previous level in the series")
        bullet("P               Invert waveform polarity  (do this before any other step)")
        bullet("N               Toggle normalised view")
        bullet("+  /  −         Scale the waveform up or down")
        blank()
        h2("   Placing peaks")
        bullet("1 – 5           Select the corresponding positive peak (P1–P5) on the "
               "current waveform")
        bullet("Shift + 1 – 5   Select the corresponding negative trough (N1–N5)")
        bullet("← / →           Move the selected peak/trough to the nearest local "
               "extremum in that direction (NZC snap)")
        bullet("Shift + ← / →   Fine-adjust the selected peak/trough by exactly one "
               "sample (bypasses snap)")
        bullet("I               Re-run full auto-detection (P1–P5 and N1–N5) for all "
               "waveforms from scratch")
        bullet("U               Propagate the currently selected peak/trough position "
               "down to all softer-level waveforms, snapping to the nearest extremum")
        blank()
        h2("   Threshold")
        bullet("Enter           Set threshold to the currently displayed waveform")
        bullet("T               Auto-estimate threshold using adjacent-level "
               "cross-covariance (Suthakar & Liberman, 2019).  "
               "The tool computes Pearson r between each pair of adjacent-level "
               "waveforms (at lag = 0), fits the r-vs-level function with a sigmoid "
               "and a power-law model, and returns the level where the better-fitting "
               "curve crosses the criterion r = 0.35.  "
               "Falls back to setting threshold at the current level if the fit fails "
               "or too few levels are available.")
        blank()
        h2("   Saving & misc.")
        bullet("S               Save peak amplitudes and latencies to *-analyzed.txt "
               "and advance to the next file in the queue")
        bullet("X               Clear the current analysis and re-run auto-detection")
        bullet("R               Restore the last saved analysis from disk")
        bullet("Ctrl + Z        Undo the last peak/threshold edit")
        blank()

        # ── 7. ABR BEST PRACTICES ─────────────────────────────────────
        h2("7.  ABR Analysis Best Practices")
        body("Before you start:  verify that all waveforms have the correct polarity. "
             "Press P to invert if the first large peak is negative rather than positive. "
             "Do this before placing any peaks — polarity affects all subsequent steps.")
        body("Recommended order of operations for each recording:")
        bullet("Step 1 — Navigate to the highest-level waveform (usually 90 dB SPL). "
               "Peaks are largest and easiest to identify here.")
        bullet("Step 2 — Correct P1–P5 from high to low level. "
               "The algorithm detects local maxima using negative zero crossings (NZCs) "
               "of the first derivative; it starts with a good guess but always verify. "
               "Select a peak with 1–5, then use ← / → to snap it to the correct "
               "extremum. Use U to propagate a corrected position to lower levels.")
        bullet("Step 3 — Once P1–P5 are set, press I to auto-estimate N1–N5 for all "
               "waveforms. The algorithm places each Nn in the trough that follows the "
               "corresponding Pn. Review and correct as needed.")
        bullet("Step 4 — Set the threshold. Press T to auto-estimate: the tool "
               "computes the normalised cross-covariance between each pair of "
               "adjacent-level waveforms, fits the resulting r-vs-level curve with "
               "a sigmoid and power-law model, and identifies the level where the "
               "best fit crosses r = 0.35 (Suthakar & Liberman, 2019). "
               "The title bar shows '✦ auto-threshold: XX dB'. Review the estimate "
               "and press Enter at the correct level to confirm or override. "
               "Any waveform below threshold will have its latencies reported as "
               "negative in the output file — this flags sub-threshold data and "
               "excludes it from Wave Growth and Latency tables.")
        bullet("Step 5 — Press S to save and move to the next file.")
        note("Wave 1 amplitude is always computed as P1 − N1. Only levels where both "
             "latencies are positive (i.e. above threshold) are included in growth "
             "function tables and plots.")
        blank()

        # ── 8. DPOAE BEST PRACTICES ───────────────────────────────────
        h2("8.  DPOAE Best Practices")
        body("DPOAE thresholds are determined by two criteria applied simultaneously:")
        bullet("Absolute DP level  ≥  criterion (dB SPL)  — the default criterion is "
               "0 dB SPL. Adjust it in the DPOAE Thresholds tab (range: −5 to +20 dB SPL). "
               "Use a higher criterion if your noise floor is elevated.")
        bullet("Signal-to-noise  ≥  5 dB  — the DP response must be at least 5 dB "
               "above the noise floor at that frequency, regardless of the absolute level. "
               "This prevents elevated noise from being misidentified as a response.")
        body("NaN (No Response) is reported when neither criterion is met at any level. "
             "NM (Not Measured) means no DP file was found for that animal / frequency.")
        note("The DPOAE audiogram excludes NaN and NM data points; only confirmed "
             "threshold values are plotted.")
        blank()

        # ── ACKNOWLEDGEMENTS ──────────────────────────────────────
        h2("9.  Acknowledgements")
        body("The peak detection algorithm and waveform analysis approach used in this "
             "tool were inspired by the EPL ABR Peak Analysis software developed by "
             "Brad Buran and Ken Hancock at the Eaton-Peabody Laboratories, "
             "Massachusetts Eye and Ear. "
             "The correlation-based threshold estimation is based on the method "
             "described by Suthakar & Liberman (2019).")
        blank()

        # ── SUMMARY HEADER (dynamic section populated by _update_overview) ──
        T.insert(tk.END, '  Loaded data summary  \n', 'sumhdr')
        T.insert(tk.END, '(no data loaded yet — click ➕ Add Data to begin)\n', 'sumtxt')

        T.config(state=tk.DISABLED)

    # ── Tab 1: ABR Thresholds ─────────────────────────────────

    def _build_threshold_tab(self):
        from matplotlib.figure import Figure
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="  ABR Thresholds  ")

        # ── Controls bar ──────────────────────────────────────────────
        top = tk.Frame(frame, pady=6)
        top.pack(fill=tk.X, padx=8)
        tk.Button(top, text="Refresh", command=self.refresh_threshold_table,
                  bg=BTN_BLUE, fg='white', padx=8).pack(side=tk.LEFT)
        tk.Button(top, text="Export to Excel", command=self.export_thresholds_excel,
                  bg=BTN_GREEN, fg='white', padx=8).pack(side=tk.LEFT, padx=6)
        tk.Button(top, text="Save Plot", command=self._save_thresh_plot,
                  bg=BTN_PURPLE, fg='white', padx=8).pack(side=tk.LEFT)

        tk.Label(top, text="  Plot:").pack(side=tk.LEFT, padx=(16, 2))
        self._thresh_plot_mode = tk.StringVar(value='mean')
        tk.Radiobutton(top, text="Mean \u00b1 SEM", variable=self._thresh_plot_mode,
                       value='mean',
                       command=self._refresh_thresh_plot).pack(side=tk.LEFT)
        tk.Radiobutton(top, text="Individual", variable=self._thresh_plot_mode,
                       value='individual',
                       command=self._refresh_thresh_plot).pack(side=tk.LEFT)

        tk.Label(top,
                 text="  NM = not measured  |  Dbl-click Group to edit  |  Dbl-click freq to re-analyze",
                 font=('Arial', 9, 'italic'), fg='gray').pack(side=tk.LEFT, padx=10)

        # ── Body: table left, audiogram right ─────────────────────────
        paned = tk.PanedWindow(frame, orient=tk.HORIZONTAL,
                               sashwidth=5, sashrelief=tk.FLAT)
        paned.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 8))

        left_frame = tk.Frame(paned)
        paned.add(left_frame, minsize=260, width=480)

        right_frame = tk.Frame(paned)
        paned.add(right_frame, minsize=300, width=480)

        def _equalise_pane(_event=None):
            total = paned.winfo_width()
            if total > 100:
                paned.sash_place(0, total // 2, 0)
                paned.unbind('<Configure>')
        paned.bind('<Configure>', _equalise_pane)

        self.thresh_tree, _ = self._make_treeview(left_frame)
        self.thresh_tree.bind('<Double-1>', self._on_thresh_dclick)

        self._thresh_fig    = Figure(figsize=(5, 4), tight_layout=True)
        self._thresh_ax     = self._thresh_fig.add_subplot(111)
        self._thresh_canvas = FigureCanvasTkAgg(self._thresh_fig, master=right_frame)
        self._thresh_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        self._thresh_plot_data = {}

    # ── Tab 2: Wave 1 Growth ──────────────────────────────────

    def _build_wave1_tab(self):
        from matplotlib.figure import Figure
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="  Wave Growth  ")

        # ── Controls bar ──────────────────────────────────────────────
        top = tk.Frame(frame, pady=6)
        top.pack(fill=tk.X, padx=8)
        tk.Button(top, text="Refresh", command=self.refresh_wave1_table,
                  bg=BTN_BLUE, fg='white', padx=8).pack(side=tk.LEFT)
        tk.Button(top, text="Export to Excel (all freqs)",
                  command=self.export_wave1_excel,
                  bg=BTN_GREEN, fg='white', padx=8).pack(side=tk.LEFT, padx=6)

        tk.Label(top, text="Frequency (kHz):").pack(side=tk.LEFT, padx=(16, 4))
        self.wave1_freq_var = tk.StringVar()
        self.wave1_freq_combo = ttk.Combobox(top, textvariable=self.wave1_freq_var,
                                              state='readonly', width=10)
        self.wave1_freq_combo.pack(side=tk.LEFT)
        self.wave1_freq_combo.bind('<<ComboboxSelected>>',
                                   lambda _e: self.refresh_wave1_table())

        tk.Label(top, text="  Wave:").pack(side=tk.LEFT, padx=(16, 2))
        self._wave_num_var = tk.StringVar(value='1')
        wave_spin = tk.Spinbox(top, from_=1, to=5, width=3,
                               textvariable=self._wave_num_var,
                               command=self.refresh_wave1_table,
                               state='readonly')
        wave_spin.pack(side=tk.LEFT)

        tk.Button(top, text="Save Plot", command=self._save_wave1_plot,
                  bg=BTN_PURPLE, fg='white', padx=8).pack(side=tk.LEFT, padx=(10, 0))

        # Plot mode toggle
        tk.Label(top, text="  Plot:").pack(side=tk.LEFT, padx=(16, 2))
        self._wave1_plot_mode = tk.StringVar(value='mean')
        tk.Radiobutton(top, text="Mean \u00b1 SEM", variable=self._wave1_plot_mode,
                       value='mean',
                       command=self._refresh_wave1_plot).pack(side=tk.LEFT)
        tk.Radiobutton(top, text="Individual", variable=self._wave1_plot_mode,
                       value='individual',
                       command=self._refresh_wave1_plot).pack(side=tk.LEFT)

        tk.Label(top, text="  Double-click a row to re-analyze",
                 font=('Arial', 9, 'italic'), fg='gray').pack(side=tk.LEFT, padx=10)

        # ── Body: table left, plot right ──────────────────────────────
        paned = tk.PanedWindow(frame, orient=tk.HORIZONTAL,
                               sashwidth=5, sashrelief=tk.FLAT)
        paned.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 8))

        left_frame = tk.Frame(paned)
        paned.add(left_frame, minsize=260, width=480)

        right_frame = tk.Frame(paned)
        paned.add(right_frame, minsize=300, width=480)

        # After layout is finalised, centre the sash so both halves are equal
        def _equalise_pane(_event=None):
            total = paned.winfo_width()
            if total > 100:
                paned.sash_place(0, total // 2, 0)
                paned.unbind('<Configure>')
        paned.bind('<Configure>', _equalise_pane)

        self.wave1_tree, _ = self._make_treeview(left_frame)
        self.wave1_tree.bind('<Double-1>', self._on_wave1_dclick)

        # Matplotlib canvas
        self._wave1_fig = Figure(figsize=(5, 4), tight_layout=True)
        self._wave1_ax  = self._wave1_fig.add_subplot(111)
        self._wave1_canvas = FigureCanvasTkAgg(self._wave1_fig, master=right_frame)
        self._wave1_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        self._wave1_plot_data = {}

    # ── Tab 3: Latencies ─────────────────────────────────────

    def _build_latency_tab(self):
        from matplotlib.figure import Figure
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="  Latencies  ")

        # ── Controls bar ──────────────────────────────────────────────
        top = tk.Frame(frame, pady=6)
        top.pack(fill=tk.X, padx=8)

        tk.Button(top, text="Refresh", command=self.refresh_latency_table,
                  bg=BTN_BLUE, fg='white', padx=8).pack(side=tk.LEFT)
        tk.Button(top, text="Export to Excel",
                  command=self.export_latency_excel,
                  bg=BTN_GREEN, fg='white', padx=8).pack(side=tk.LEFT, padx=6)

        tk.Label(top, text="Frequency (kHz):").pack(side=tk.LEFT, padx=(16, 4))
        self.lat_freq_var = tk.StringVar()
        self.lat_freq_combo = ttk.Combobox(top, textvariable=self.lat_freq_var,
                                           state='readonly', width=10)
        self.lat_freq_combo.pack(side=tk.LEFT)
        self.lat_freq_combo.bind('<<ComboboxSelected>>',
                                 lambda _e: self.refresh_latency_table())

        tk.Label(top, text="  Wave:").pack(side=tk.LEFT, padx=(16, 2))
        self._lat_wave_var = tk.StringVar(value='1')
        tk.Spinbox(top, from_=1, to=5, width=3,
                   textvariable=self._lat_wave_var,
                   command=self.refresh_latency_table,
                   state='readonly').pack(side=tk.LEFT)

        tk.Label(top, text="  Peak:").pack(side=tk.LEFT, padx=(12, 2))
        self._lat_peak_var = tk.StringVar(value='P')
        tk.Radiobutton(top, text="P (positive)", variable=self._lat_peak_var,
                       value='P', command=self.refresh_latency_table).pack(side=tk.LEFT)
        tk.Radiobutton(top, text="N (negative)", variable=self._lat_peak_var,
                       value='N', command=self.refresh_latency_table).pack(side=tk.LEFT)

        tk.Button(top, text="Save Plot", command=self._save_latency_plot,
                  bg=BTN_PURPLE, fg='white', padx=8).pack(side=tk.LEFT, padx=(10, 0))

        # Plot mode toggle
        tk.Label(top, text="  Plot:").pack(side=tk.LEFT, padx=(16, 2))
        self._lat_plot_mode = tk.StringVar(value='mean')
        tk.Radiobutton(top, text="Mean \u00b1 SEM", variable=self._lat_plot_mode,
                       value='mean', command=self._refresh_latency_plot).pack(side=tk.LEFT)
        tk.Radiobutton(top, text="Individual", variable=self._lat_plot_mode,
                       value='individual', command=self._refresh_latency_plot).pack(side=tk.LEFT)

        tk.Label(top, text="  Double-click a row to re-analyze",
                 font=('Arial', 9, 'italic'), fg='gray').pack(side=tk.LEFT, padx=10)

        # ── Body: table left, plot right ──────────────────────────────
        paned = tk.PanedWindow(frame, orient=tk.HORIZONTAL,
                               sashwidth=5, sashrelief=tk.FLAT)
        paned.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 8))

        left_frame = tk.Frame(paned)
        paned.add(left_frame, minsize=260, width=480)

        right_frame = tk.Frame(paned)
        paned.add(right_frame, minsize=300, width=480)

        def _equalise(_event=None):
            total = paned.winfo_width()
            if total > 100:
                paned.sash_place(0, total // 2, 0)
                paned.unbind('<Configure>')
        paned.bind('<Configure>', _equalise)

        self.lat_tree, _ = self._make_treeview(left_frame)
        self.lat_tree.bind('<Double-1>', self._on_lat_dclick)

        self._lat_fig    = Figure(figsize=(5, 4), tight_layout=True)
        self._lat_ax     = self._lat_fig.add_subplot(111)
        self._lat_canvas = FigureCanvasTkAgg(self._lat_fig, master=right_frame)
        self._lat_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        self._lat_plot_data = {}   # {mid: {level: latency_ms}}
        self._lat_raw_map   = {}   # {mid: raw_filepath}

    def refresh_latency_table(self):
        if not self.mice_data:
            return
        try:
            target_freq = float(self.lat_freq_var.get())
        except (ValueError, tk.TclError):
            return

        try:
            wn = int(self._lat_wave_var.get())
        except (ValueError, tk.TclError):
            wn = 1
        peak    = self._lat_peak_var.get()   # 'P' or 'N'
        lat_col = f'{peak}{wn} Latency'

        lat_dict   = {}   # {mid: {level: latency_ms}}
        all_levels = set()
        self._lat_raw_map = {}

        for mid, mdata in self.mice_data.items():
            if not mdata['analyzed_files'] and not mdata['raw_abr_files']:
                continue
            for fp in mdata['analyzed_files']:
                p = self._get_analyzed(fp)
                if p['frequency'] is None:
                    continue
                if abs(p['frequency'] - target_freq) > 0.1:
                    continue
                if p['data'] is None:
                    continue

                df = p['data']
                lat_dict[mid] = {}
                self._lat_raw_map[mid] = fp.replace('-analyzed.txt', '')

                for _, row in df.iterrows():
                    level = float(row['Level'])
                    lat   = row.get(lat_col, np.nan)
                    all_levels.add(level)   # track every tested level
                    if not np.isnan(lat) and lat > 0:
                        lat_dict[mid][level] = round(float(lat), 4)
                    else:
                        lat_dict[mid][level] = np.nan  # sub-threshold

        if not lat_dict:
            return

        levels = sorted(all_levels, reverse=True)
        cols   = ['Mouse', 'Group'] + [f"{l:.0f} dB" for l in levels]
        self._configure_tree(self.lat_tree, cols)

        for mid in sorted(lat_dict.keys()):
            group = self._get_group(mid)
            row   = [mid, group]
            for lv in levels:
                if lv not in lat_dict[mid]:
                    row.append('NM')
                elif np.isnan(lat_dict[mid][lv]):
                    row.append('NaN')
                else:
                    row.append(f"{lat_dict[mid][lv]:.4f}")
            self.lat_tree.insert('', tk.END, values=row)

        self._lat_plot_data = lat_dict
        self._refresh_latency_plot()

    def _refresh_latency_plot(self):
        ax = self._lat_ax
        ax.clear()

        data = getattr(self, '_lat_plot_data', {})
        if not data:
            ax.text(0.5, 0.5, 'No data — click Refresh',
                    ha='center', va='center', transform=ax.transAxes, color='gray')
            self._lat_canvas.draw()
            return

        mode = self._lat_plot_mode.get()
        all_levels = sorted({lv for d in data.values() for lv in d})

        palette = ['#2196F3', '#E53935', '#43A047', '#FB8C00',
                   '#8E24AA', '#00ACC1', '#6D4C41', '#FFB300']
        group_data = {}
        for mid, d in data.items():
            grp = self._get_group(mid)
            if not grp.strip() or grp.strip().upper() == 'N/A':  # excluded from plot
                continue
            group_data.setdefault(grp, {})[mid] = d

        if mode == 'mean':
            has_real = any(len(m) > 1 for m in group_data.values())
            if not has_real and len(group_data) > 1:
                group_data = {'All Mice': {m: d for m, d in data.items()}}

        group_names = sorted(group_data.keys())
        color_map   = {g: palette[i % len(palette)] for i, g in enumerate(group_names)}

        if mode == 'mean':
            for grp in group_names:
                mice_dicts = list(group_data[grp].values())
                means, sems, valid_lvs = [], [], []
                for lv in all_levels:
                    vals = [d[lv] for d in mice_dicts if lv in d]
                    if not vals:
                        continue
                    m = float(np.nanmean(vals))
                    s = (float(np.nanstd(vals, ddof=1)) / np.sqrt(np.sum(~np.isnan(vals)))
                         if len(vals) > 1 else 0.0)
                    means.append(m); sems.append(s); valid_lvs.append(lv)
                if not valid_lvs:
                    continue
                c = color_map[grp]
                n = len(mice_dicts)
                ax.plot(valid_lvs, means, '-o', color=c,
                        label=f"{grp} (n={n})", linewidth=2, markersize=5)
                ax.fill_between(valid_lvs,
                                [m - s for m, s in zip(means, sems)],
                                [m + s for m, s in zip(means, sems)],
                                alpha=0.2, color=c)

        else:  # individual — thin traces + thick group mean
            for grp in group_names:
                c          = color_map[grp]
                mice_dicts = list(group_data[grp].values())
                for mid, d in sorted(group_data[grp].items()):
                    lvs  = sorted(d.keys())
                    lats = [d[lv] for lv in lvs]
                    ax.plot(lvs, lats, '-', color=c, alpha=0.3, linewidth=1.0)
                means, valid_lvs = [], []
                for lv in all_levels:
                    vals = [d[lv] for d in mice_dicts if lv in d]
                    if not vals:
                        continue
                    means.append(float(np.nanmean(vals)))
                    valid_lvs.append(lv)
                if valid_lvs:
                    n = len(mice_dicts)
                    ax.plot(valid_lvs, means, '-o', color=c, linewidth=2.5,
                            markersize=6, label=f"{grp} mean (n={n})", zorder=5)

        try:
            freq = float(self.lat_freq_var.get())
            wn   = int(self._lat_wave_var.get())
            peak = self._lat_peak_var.get()
            ax.set_title(f"Wave {wn} {peak}-peak Latency  —  {freq:.2f} kHz",
                         fontsize=10)
        except (ValueError, AttributeError):
            ax.set_title("Peak Latency – Intensity Function", fontsize=10)

        ax.set_xlabel("Level (dB SPL)", fontsize=9)
        ax.set_ylabel("Latency (ms)", fontsize=9)
        ax.grid(True, alpha=0.3, linestyle='--')
        ax.tick_params(labelsize=8)

        handles, labels = ax.get_legend_handles_labels()
        if handles:
            ax.legend(fontsize=8, ncol=2 if len(handles) > 8 else 1, framealpha=0.8)

        self._lat_fig.tight_layout()
        self._lat_canvas.draw()

    def _on_lat_dclick(self, event):
        row_id = self.lat_tree.identify_row(event.y)
        if not row_id:
            return
        vals = self.lat_tree.item(row_id, 'values')
        if not vals:
            return
        mid = vals[0]
        raw_path = self._lat_raw_map.get(mid)
        if not raw_path:
            messagebox.showwarning("No raw file", f"No raw ABR file found for {mid}.")
            return
        self._open_peak_analysis([(mid, raw_path)])

    def _save_latency_plot(self):
        fp = filedialog.asksaveasfilename(
            defaultextension='.png',
            filetypes=[('PNG image', '*.png'), ('PDF', '*.pdf'), ('SVG', '*.svg')],
            title="Save Latency Plot")
        if fp:
            self._lat_fig.savefig(fp, dpi=150, bbox_inches='tight')
            messagebox.showinfo("Saved", f"Figure saved:\n{fp}")

    def export_latency_excel(self):
        if not self.mice_data:
            messagebox.showwarning("No data", "Load a data folder first.")
            return
        fp = filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            filetypes=[('Excel workbook', '*.xlsx')],
            title="Save Latency Data")
        if not fp:
            return
        try:
            wb = openpyxl.Workbook()
            first = True

            all_freqs = set()
            for mdata in self.mice_data.values():
                for afp in mdata['analyzed_files']:
                    p = self._get_analyzed(afp)
                    if p['frequency'] is not None:
                        all_freqs.add(p['frequency'])

            for freq in sorted(all_freqs):
                self.lat_freq_var.set(f"{freq:.2f}")
                # Export one sheet per (freq, wave, peak) combination
                for wn in range(1, 6):
                    for peak in ('P', 'N'):
                        self._lat_wave_var.set(str(wn))
                        self._lat_peak_var.set(peak)
                        self.refresh_latency_table()
                        df = self._tree_to_df(self.lat_tree)
                        if df.empty:
                            continue
                        sheet_name = f"Lat_{freq:.1f}kHz_W{wn}{peak}"[:31]
                        if first:
                            ws = wb.active
                            ws.title = sheet_name
                            first = False
                        else:
                            ws = wb.create_sheet(title=sheet_name)
                        self._write_df_to_sheet(
                            ws, df,
                            f"Latency (ms)  —  {freq:.2f} kHz  Wave {wn} {peak}-peak")

            wb.save(fp)
            messagebox.showinfo("Saved", f"Saved:\n{fp}")
        except Exception as exc:
            messagebox.showerror("Export error", str(exc))

    # ── Tab 4: Plot Traces ────────────────────────────────────

    def _build_plot_tab(self):
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="  Plot Traces  ")

        # Left control panel
        ctrl = tk.Frame(frame, width=210, relief=tk.RIDGE, bd=1)
        ctrl.pack(side=tk.LEFT, fill=tk.Y, padx=6, pady=6)
        ctrl.pack_propagate(False)

        tk.Label(ctrl, text="Plot Controls",
                 font=('Arial', 10, 'bold')).pack(pady=(10, 4))

        tk.Label(ctrl, text="Select Mice:", anchor='w').pack(anchor='w', padx=6)
        lf = tk.Frame(ctrl)
        lf.pack(fill=tk.BOTH, padx=6, expand=True)
        self.mice_listbox = tk.Listbox(lf, selectmode=tk.MULTIPLE,
                                        height=10, exportselection=False,
                                        font=('Courier', 9))
        msb = ttk.Scrollbar(lf, command=self.mice_listbox.yview)
        self.mice_listbox.configure(yscrollcommand=msb.set)
        msb.pack(side=tk.RIGHT, fill=tk.Y)
        self.mice_listbox.pack(fill=tk.BOTH, expand=True)

        br = tk.Frame(ctrl)
        br.pack(fill=tk.X, padx=6, pady=(2, 4))
        tk.Button(br, text="All",
                  command=lambda: self.mice_listbox.select_set(0, tk.END),
                  width=7).pack(side=tk.LEFT)
        tk.Button(br, text="None",
                  command=lambda: self.mice_listbox.selection_clear(0, tk.END),
                  width=7).pack(side=tk.LEFT, padx=4)

        tk.Label(ctrl, text="Frequency (kHz):", anchor='w').pack(anchor='w', padx=6,
                                                                   pady=(4, 0))
        self.plot_freq_var = tk.StringVar()
        self.plot_freq_combo = ttk.Combobox(ctrl, textvariable=self.plot_freq_var,
                                             state='readonly', width=14)
        self.plot_freq_combo.pack(padx=6, fill=tk.X)
        self.plot_freq_combo.bind('<<ComboboxSelected>>',
                                  lambda _e: self._update_levels_listbox())

        tk.Label(ctrl, text="Levels to overlay (dB SPL):",
                 anchor='w').pack(anchor='w', padx=6, pady=(8, 0))
        tk.Label(ctrl, text="Ctrl+click to multi-select",
                 font=('Arial', 8), fg='gray').pack(anchor='w', padx=6)
        self.levels_listbox = tk.Listbox(ctrl, selectmode=tk.MULTIPLE, height=9,
                                          exportselection=False, font=('Courier', 9))
        self.levels_listbox.pack(fill=tk.X, padx=6)

        # Layout / content toggle (vertical list for readability)
        self.plot_mode_var = tk.StringVar(value='stacked')
        tk.Label(ctrl, text="Layout:", anchor='w').pack(anchor='w', padx=6, pady=(8, 0))
        mode_frame = tk.Frame(ctrl)
        mode_frame.pack(fill=tk.X, padx=6)
        for _lbl, _val in [
            ("Stacked (mean \u00b1 SEM)",          'stacked'),
            ("Overlay (mean \u00b1 SEM)",           'overlay'),
            ("By Group (mean \u00b1 SEM)",          'bygroup'),
            ("Mean + Individuals",                  'individuals'),
        ]:
            ttk.Radiobutton(mode_frame, text=_lbl,
                            variable=self.plot_mode_var,
                            value=_val).pack(anchor='w')

        tk.Button(ctrl, text="Plot", command=self.plot_traces,
                  bg=BTN_BLUE, fg='white', pady=4).pack(pady=(10, 4), padx=6, fill=tk.X)
        tk.Button(ctrl, text="Save Figure", command=self.save_plot_figure,
                  bg=BTN_PURPLE, fg='white', pady=4).pack(padx=6, fill=tk.X)

        # Right: matplotlib canvas
        plot_frame = tk.Frame(frame)
        plot_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.plot_fig = Figure(figsize=(7, 5), dpi=100)
        self.plot_canvas = FigureCanvasTkAgg(self.plot_fig, master=plot_frame)
        self.plot_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        toolbar_frame = tk.Frame(plot_frame)
        toolbar_frame.pack(fill=tk.X)
        NavigationToolbar2Tk(self.plot_canvas, toolbar_frame)

    # ── Tab 4: DPOAE Thresholds ───────────────────────────────

    def _build_dpoae_tab(self):
        from matplotlib.figure import Figure
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="  DPOAE Thresholds  ")

        # ── Controls bar ──────────────────────────────────────────────
        top = tk.Frame(frame, pady=6)
        top.pack(fill=tk.X, padx=8)
        tk.Label(top, text="Criterion (dB SPL):").pack(side=tk.LEFT)
        self.dpoae_criterion = tk.DoubleVar(value=0.0)
        tk.Spinbox(top, from_=-5, to=20, increment=1,
                   textvariable=self.dpoae_criterion,
                   width=6, font=('Arial', 10)).pack(side=tk.LEFT, padx=4)
        tk.Button(top, text="Calculate", command=self.refresh_dpoae_table,
                  bg=BTN_BLUE, fg='white', padx=8).pack(side=tk.LEFT, padx=6)
        tk.Button(top, text="Export to Excel", command=self.export_dpoae_excel,
                  bg=BTN_GREEN, fg='white', padx=8).pack(side=tk.LEFT)
        tk.Button(top, text="Save Plot", command=self._save_dpoae_plot,
                  bg=BTN_PURPLE, fg='white', padx=8).pack(side=tk.LEFT, padx=6)

        # Plot mode toggle
        tk.Label(top, text="  Plot:").pack(side=tk.LEFT, padx=(16, 2))
        self._dpoae_plot_mode = tk.StringVar(value='mean')
        tk.Radiobutton(top, text="Mean \u00b1 SEM", variable=self._dpoae_plot_mode,
                       value='mean',
                       command=self._refresh_dpoae_plot).pack(side=tk.LEFT)
        tk.Radiobutton(top, text="Individual", variable=self._dpoae_plot_mode,
                       value='individual',
                       command=self._refresh_dpoae_plot).pack(side=tk.LEFT)

        tk.Label(top,
                 text="  Threshold = lowest level \u2265 criterion AND \u2265 5 dB above noise floor  |  NaN = no response",
                 font=('Arial', 9, 'italic'), fg='gray').pack(side=tk.LEFT, padx=8)

        # ── Body: table left, plot right ──────────────────────────────
        paned = tk.PanedWindow(frame, orient=tk.HORIZONTAL,
                               sashwidth=5, sashrelief=tk.FLAT)
        paned.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 8))

        left_frame = tk.Frame(paned)
        paned.add(left_frame, minsize=260, width=480)

        right_frame = tk.Frame(paned)
        paned.add(right_frame, minsize=300, width=480)

        def _equalise(_event=None):
            total = paned.winfo_width()
            if total > 100:
                paned.sash_place(0, total // 2, 0)
                paned.unbind('<Configure>')
        paned.bind('<Configure>', _equalise)

        self.dpoae_tree, _ = self._make_treeview(left_frame)

        self._dpoae_fig    = Figure(figsize=(5, 4), tight_layout=True)
        self._dpoae_ax     = self._dpoae_fig.add_subplot(111)
        self._dpoae_canvas = FigureCanvasTkAgg(self._dpoae_fig, master=right_frame)
        self._dpoae_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        self._dpoae_plot_data = {}

    # ----------------------------------------------------------
    # Helper: reusable scrollable Treeview
    # ----------------------------------------------------------

    def _make_treeview(self, parent):
        outer = tk.Frame(parent)
        outer.pack(fill=tk.BOTH, expand=True, padx=8, pady=4)
        tree = ttk.Treeview(outer, show='headings')
        vsb  = ttk.Scrollbar(outer, orient='vertical',   command=tree.yview)
        hsb  = ttk.Scrollbar(outer, orient='horizontal', command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        hsb.pack(side=tk.BOTTOM, fill=tk.X)
        vsb.pack(side=tk.RIGHT,  fill=tk.Y)
        tree.pack(fill=tk.BOTH, expand=True)
        return tree, outer

    # ===========================================================
    # Data Loading
    # ===========================================================

    def _add_data(self):
        """Single entry point: ask the user whether to pick folders or files."""
        dlg = tk.Toplevel(self.root)
        dlg.title("Add Data")
        dlg.geometry("300x140")
        dlg.resizable(False, False)
        dlg.grab_set()

        tk.Label(dlg, text="How would you like to add data?",
                 font=('Arial', 10)).pack(pady=(18, 10))

        btn_row = tk.Frame(dlg)
        btn_row.pack()

        def _pick_folder():
            dlg.destroy()
            self.load_data_folder()

        def _pick_files():
            dlg.destroy()
            self._open_files_direct()

        _mk_btn(btn_row, text="📂  Add Folder(s)", command=_pick_folder,
                bg='#3498db', width=16, pady=4).pack(side=tk.LEFT, padx=8)
        _mk_btn(btn_row, text="📄  Open Files", command=_pick_files,
                bg='#27ae60', width=16, pady=4).pack(side=tk.LEFT)

        tk.Button(dlg, text="Cancel", command=dlg.destroy,
                  pady=2).pack(pady=(12, 0))

    def load_data_folder(self):
        folders = _pick_folders(self.root, title="Select one or more data folders")
        if not folders:
            return

        all_new_mice = {}
        skipped = []
        for folder in folders:
            found = find_mouse_data(folder)
            if not found:
                skipped.append(folder)
            else:
                all_new_mice.update(found)

        if skipped:
            messagebox.showwarning("Some folders skipped",
                "No data found in:\n" + "\n".join(f"  {f}" for f in skipped))

        if not all_new_mice:
            messagebox.showwarning("No data found",
                "No *-analyzed.txt or DP-* files were found in the selected folder(s).\n"
                "Make sure you selected the correct root directory.")
            return

        duplicates = [mid for mid in all_new_mice if mid in self.mice_data]
        if duplicates:
            answer = messagebox.askyesno(
                "Duplicate mouse IDs",
                f"The following mouse ID(s) are already loaded and will be replaced:\n\n"
                f"  {', '.join(duplicates)}\n\nContinue?")
            if not answer:
                return

        label = folders[0] if len(folders) == 1 else f"{len(folders)} folders"
        self._merge_and_refresh(all_new_mice, label=label)

    def _merge_and_refresh(self, new_mice, label=""):
        folder = label  # used in status label below

        # Merge — new entries overwrite duplicates
        self.mice_data.update(new_mice)

        # Rebuild frequency lists from ALL loaded mice
        freqs = set()
        for mdata in self.mice_data.values():
            for fp in mdata['analyzed_files']:
                parsed = self._get_analyzed(fp)
                if parsed['frequency'] is not None:
                    freqs.add(parsed['frequency'])
        freq_strs = [f"{f:.2f}" for f in sorted(freqs)]

        self.wave1_freq_combo['values'] = freq_strs
        if freq_strs and self.wave1_freq_var.get() not in freq_strs:
            self.wave1_freq_var.set(freq_strs[0])

        self.lat_freq_combo['values'] = freq_strs
        if freq_strs and self.lat_freq_var.get() not in freq_strs:
            self.lat_freq_var.set(freq_strs[0])

        self.plot_freq_combo['values'] = freq_strs
        if freq_strs and self.plot_freq_var.get() not in freq_strs:
            self.plot_freq_var.set(freq_strs[0])

        # Rebuild mice listbox (full sorted list)
        self.mice_listbox.delete(0, tk.END)
        for mid in sorted(self.mice_data.keys()):
            self.mice_listbox.insert(tk.END, mid)

        # Update status label
        n = len(self.mice_data)
        self.folder_label.config(
            text=f"{n} mouse/session{'s' if n != 1 else ''} loaded  |  last added: {folder}")

        self._update_levels_listbox()
        self._update_overview()

        # Refresh all tables
        self.refresh_threshold_table()
        self.refresh_wave1_table()
        self.refresh_dpoae_table()

        messagebox.showinfo("Folder added",
            f"Added {len(new_mice)} mouse/session(s) from:\n{folder}\n\n"
            f"Total loaded: {len(self.mice_data)} mouse/session(s)\n"
            f"ABR frequencies: {', '.join(freq_strs) if freq_strs else 'none yet'} kHz")

        # ── Check for unanalyzed raw ABR files ────────────────────
        unanalyzed_queue = []
        for mid, mdata in new_mice.items():
            for fp in mdata.get('unanalyzed_raw', []):
                unanalyzed_queue.append((mid, fp))

        if unanalyzed_queue:
            n_files = len(unanalyzed_queue)
            mice_set = sorted({mid for mid, _ in unanalyzed_queue})
            answer = messagebox.askyesno(
                "Unanalyzed ABR files found",
                f"Found {n_files} unanalyzed ABR file(s) across "
                f"{len(mice_set)} mouse/session(s):\n\n"
                f"  {', '.join(mice_set)}\n\n"
                "Would you like to open the Peak Analysis tool now\n"
                "to set thresholds and label peaks?")
            if answer:
                self._open_peak_analysis(unanalyzed_queue)

    def _open_files_direct(self):
        """Let the user pick any ABR files directly and open them for peak analysis."""
        paths = filedialog.askopenfilenames(
            title="Select ABR files for Peak Analysis",
            filetypes=[
                ("All ABR files", "*.tsv *.txt *.csv *"),
                ("CFTS3 FAST ABR (.tsv)", "*.tsv"),
                ("Text ABR (.txt)", "*.txt"),
                ("CSV ABR (.csv)", "*.csv"),
                ("CFTS extensionless", "*"),
            ])
        if not paths:
            return

        # Separate analyzed files from raw files — analyzed ones belong in Add Folder
        analyzed_selected = [os.path.basename(fp) for fp in paths
                              if fp.endswith('-analyzed.txt')]
        raw_paths = [fp for fp in paths if not fp.endswith('-analyzed.txt')]

        if analyzed_selected:
            messagebox.showinfo(
                "Use 'Add Folder' for analyzed files",
                "The following file(s) are already-analyzed results.\n"
                "Use 📂 Add Folder to load their threshold data instead:\n\n" +
                "\n".join(f"  {f}" for f in analyzed_selected))

        queue = []
        skipped = []
        for fp in raw_paths:
            raw = parse_raw_abr_file(fp)
            if raw['waveforms'] is None or not raw['levels']:
                skipped.append(os.path.basename(fp))
            else:
                mouse_id = os.path.splitext(os.path.basename(fp))[0]
                queue.append((mouse_id, fp))

        if skipped:
            messagebox.showwarning(
                "Some files skipped",
                "Could not parse the following files:\n" +
                "\n".join(f"  {f}" for f in skipped))

        if queue:
            self._open_peak_analysis(queue)

    def _open_peak_analysis(self, file_queue):
        """
        Open (or re-open) the PeakAnalysisWindow for the given queue.
        The on_complete callback refreshes the main window tables and
        adds any newly created -analyzed.txt files to the mouse record.
        """
        def _on_file_saved(mouse_id, analyzed_path):
            """Called by the window each time a file is saved.

            Updates mice_data, invalidates the parse cache, then immediately
            refreshes all tables so the Wave 1 / Threshold views show the new
            peak data without waiting for the analysis window to close.
            """
            # 1. Keep mice_data in sync
            if mouse_id in self.mice_data:
                existing = self.mice_data[mouse_id]['analyzed_files']
                if analyzed_path not in existing:
                    existing.append(analyzed_path)
                    existing.sort()
                # Remove from unanalyzed list
                raw_path = analyzed_path.replace('-analyzed.txt', '')
                unanalyzed = self.mice_data[mouse_id].get('unanalyzed_raw', [])
                if raw_path in unanalyzed:
                    unanalyzed.remove(raw_path)

            # 2. Bust the cache so the re-parsed file is used on next read
            self.analyzed_cache.pop(analyzed_path, None)

            # 3. Refresh tables immediately — do not wait for window close
            self.refresh_threshold_table()
            self.refresh_wave1_table()

        win = PeakAnalysisWindow(self.root, file_queue,
                                 on_complete=_on_file_saved)
        # Also do a full refresh (freq combos, mice list, etc.) when the
        # entire queue finishes and the window closes.
        _orig_finish = win._finish
        def _patched_finish():
            self._refresh_all()
            _orig_finish()
        win._finish = _patched_finish

    def _refresh_all(self):
        """Rebuild frequency combos, mice list, and all tables."""
        freqs = set()
        for mdata in self.mice_data.values():
            for fp in mdata['analyzed_files']:
                p = self._get_analyzed(fp)
                if p['frequency'] is not None:
                    freqs.add(p['frequency'])
        freq_strs = [f"{f:.2f}" for f in sorted(freqs)]

        self.wave1_freq_combo['values'] = freq_strs
        if freq_strs and self.wave1_freq_var.get() not in freq_strs:
            self.wave1_freq_var.set(freq_strs[0])
        self.lat_freq_combo['values'] = freq_strs
        if freq_strs and self.lat_freq_var.get() not in freq_strs:
            self.lat_freq_var.set(freq_strs[0])
        self.plot_freq_combo['values'] = freq_strs
        if freq_strs and self.plot_freq_var.get() not in freq_strs:
            self.plot_freq_var.set(freq_strs[0])

        self.mice_listbox.delete(0, tk.END)
        for mid in sorted(self.mice_data.keys()):
            self.mice_listbox.insert(tk.END, mid)

        self._update_levels_listbox()
        self._update_overview()
        self.refresh_threshold_table()
        self.refresh_wave1_table()
        self.refresh_dpoae_table()

    def remove_animals(self):
        """Remove whichever sessions are currently selected in the active tab's table."""
        # Map tab index → treeview
        tab_trees = {
            1: self.thresh_tree,
            2: self.wave1_tree,
            3: self.lat_tree,
            5: self.dpoae_tree,
        }
        try:
            tab_idx = self.notebook.index(self.notebook.select())
        except tk.TclError:
            tab_idx = -1

        tree = tab_trees.get(tab_idx)
        if tree is None:
            messagebox.showinfo(
                "Select a session",
                "Switch to a data tab (ABR Thresholds, Wave Growth, Latencies, "
                "or DPOAE Thresholds), highlight the row(s) you want to remove, "
                "then click Remove Session.")
            return

        selected_ids = tree.selection()   # tuple of iid strings
        if not selected_ids:
            messagebox.showinfo(
                "Nothing selected",
                "Click a row in the table to select a session, "
                "then click Remove Session.\n"
                "Hold Ctrl to select multiple rows.")
            return

        to_remove = []
        for iid in selected_ids:
            vals = tree.item(iid, 'values')
            if vals:
                to_remove.append(vals[0])   # column 0 = Mouse / session ID

        if not to_remove:
            return

        label = ', '.join(to_remove)
        if not messagebox.askyesno(
                "Remove session(s)?",
                f"Remove the following session(s) from the dataset?\n\n  {label}"):
            return

        for mid in to_remove:
            self.mice_data.pop(mid, None)
            self._group_overrides.pop(mid, None)
            for cache in (self.analyzed_cache, self.raw_cache, self.dp_cache):
                for k in list(cache.keys()):
                    if os.path.basename(os.path.dirname(k)) == mid:
                        cache.pop(k, None)

        self._merge_and_refresh({}, label=f"{len(self.mice_data)} session(s) loaded")

    def clear_all_data(self):
        if self.mice_data:
            if not messagebox.askyesno("Clear all data",
                    f"Remove all {len(self.mice_data)} loaded mouse/session(s) and start fresh?"):
                return
        self.mice_data       = {}
        self.analyzed_cache  = {}
        self.raw_cache       = {}
        self.dp_cache        = {}
        self._group_overrides = {}
        self._thresh_raw_map  = {}
        self._wave1_raw_map   = {}

        self.folder_label.config(text="No data loaded")
        self.wave1_freq_combo['values'] = []
        self.wave1_freq_var.set('')
        self.lat_freq_combo['values']   = []
        self.lat_freq_var.set('')
        self.plot_freq_combo['values']  = []
        self.plot_freq_var.set('')
        self.mice_listbox.delete(0, tk.END)
        self.levels_listbox.delete(0, tk.END)

        for tree in (self.thresh_tree, self.wave1_tree, self.lat_tree, self.dpoae_tree):
            tree['columns'] = []
            tree.delete(*tree.get_children())

        self.plot_fig.clear()
        self.plot_canvas.draw()
        self._update_overview()

    def _on_tab_change(self, _event=None):
        """Auto-refresh the newly selected tab if data is loaded."""
        if not self.mice_data:
            return
        try:
            idx  = self.notebook.index(self.notebook.select())
            name = self._tab_names[idx]
            fn   = self._tab_refresh.get(name)
            if fn:
                fn()
        except (tk.TclError, IndexError, AttributeError):
            pass

    def _update_overview(self):
        """Replace only the live summary section at the end of the overview widget."""
        T = self.overview_text
        T.config(state=tk.NORMAL)

        # Locate the summary header tag and delete everything from it onwards
        idx = T.search('  Loaded data summary  ', '1.0', tk.END)
        if idx:
            T.delete(idx, tk.END)
        else:
            # Fallback: append after existing content
            pass

        n = len(self.mice_data)
        T.insert(tk.END, '\n', 'body')
        T.insert(tk.END, '  Loaded data summary  \n', 'sumhdr')

        if not self.mice_data:
            T.insert(tk.END,
                     '(no data loaded yet — click ➕ Add Data to begin)\n',
                     'sumtxt')
        else:
            T.insert(tk.END,
                     f'  {n} animal/session(s) loaded\n'
                     f'  {"─" * 68}\n',
                     'sumtxt')
            for mid in sorted(self.mice_data.keys()):
                md  = self.mice_data[mid]
                grp = self._get_group(mid) or '(unassigned)'
                T.insert(tk.END,
                         f'\n  {mid}   [group: {grp}]\n'
                         f'    analyzed: {len(md["analyzed_files"])}  '
                         f'raw ABR: {len(md["raw_abr_files"])}  '
                         f'DPOAE: {len(md["dp_files"])}\n',
                         'sumtxt')
                for fp in md['analyzed_files']:
                    p = self._get_analyzed(fp)
                    if p['frequency'] is not None:
                        thr = (f"{p['threshold']:.0f} dB SPL"
                               if p['threshold'] is not None else 'NaN')
                        T.insert(tk.END,
                                 f'      {os.path.basename(fp):<38}'
                                 f'{p["frequency"]:5.1f} kHz   thr = {thr}\n',
                                 'sumtxt')

        T.config(state=tk.DISABLED)
        T.see(tk.END)   # scroll to summary so user sees fresh data

    def _update_levels_listbox(self):
        """Refresh the levels list for the currently-selected plot frequency."""
        try:
            target_freq = float(self.plot_freq_var.get())
        except (ValueError, tk.TclError):
            return

        all_levels = set()
        for mdata in self.mice_data.values():
            for fp in mdata['analyzed_files']:
                p = self._get_analyzed(fp)
                if p['frequency'] is not None and abs(p['frequency'] - target_freq) < 0.1:
                    if p['data'] is not None:
                        for lv in p['data']['Level']:
                            all_levels.add(float(lv))

        self.levels_listbox.delete(0, tk.END)
        for lv in sorted(all_levels, reverse=True):
            self.levels_listbox.insert(tk.END, f"{lv:.0f}")

    # ===========================================================
    # Cache helpers
    # ===========================================================

    def _get_analyzed(self, fp):
        if fp not in self.analyzed_cache:
            self.analyzed_cache[fp] = parse_analyzed_file(fp)
        return self.analyzed_cache[fp]

    def _get_raw(self, fp):
        if fp not in self.raw_cache:
            self.raw_cache[fp] = parse_raw_abr_file(fp)
        return self.raw_cache[fp]

    def _get_dp(self, fp):
        if fp not in self.dp_cache:
            self.dp_cache[fp] = parse_dpoae_file(fp)
        return self.dp_cache[fp]

    def _get_group(self, mid):
        """Return the group label for a mouse, respecting any user override.
        Returns '' (empty string) when the animal is unassigned / null group."""
        if mid in self._group_overrides:
            return self._group_overrides[mid]
        return self.mice_data.get(mid, {}).get('group', '')

    def _is_assigned(self, mid):
        """True only when the session has a non-blank, non-N/A group label."""
        grp = self._get_group(mid).strip()
        return bool(grp) and grp.upper() != 'N/A'

    def _start_group_edit(self, tree, event):
        """
        Inline-edit the Group cell in a Treeview on double-click.
        Places a temporary Entry widget over the clicked cell.
        """
        row_id = tree.identify_row(event.y)
        col_id = tree.identify_column(event.x)
        if not row_id:
            return

        # Only allow editing the 'Group' column (column #2 = index 1)
        cols = tree['columns']
        try:
            col_idx = int(col_id.lstrip('#')) - 1
        except ValueError:
            return
        if col_idx < 0 or col_idx >= len(cols) or cols[col_idx] != 'Group':
            return

        # Mouse ID is always column 0
        values  = tree.item(row_id, 'values')
        mid     = values[0]
        current = values[1]

        # Bounding box of the cell
        bbox = tree.bbox(row_id, col_id)
        if not bbox:
            return
        x, y, w, h = bbox

        entry_var = tk.StringVar(value=current)
        entry = tk.Entry(tree, textvariable=entry_var, font=('Arial', 9))
        entry.place(x=x, y=y, width=w, height=h)
        entry.focus_set()
        entry.select_range(0, tk.END)

        def _commit(evt=None):
            new_group = entry_var.get().strip()
            entry.destroy()
            if new_group == current:
                return
            self._group_overrides[mid] = new_group
            # Refresh all tables so the change propagates everywhere
            self.refresh_threshold_table()
            self.refresh_wave1_table()
            self.refresh_dpoae_table()
            self._update_overview()

        entry.bind('<Return>',    _commit)
        entry.bind('<FocusOut>',  _commit)
        entry.bind('<Escape>',    lambda e: entry.destroy())

    def _reanalyze_from_map(self, mid, raw_path):
        """Open peak analysis for a single (mouse_id, raw_path) pair."""
        if not os.path.exists(raw_path):
            messagebox.showwarning("File not found",
                                   f"Raw ABR file not found:\n{raw_path}")
            return
        self._open_peak_analysis([(mid, raw_path)])

    def _on_thresh_dclick(self, event):
        """Double-click on threshold table: Group column → edit; freq column → re-analyze."""
        tree   = self.thresh_tree
        row_id = tree.identify_row(event.y)
        col_id = tree.identify_column(event.x)
        if not row_id:
            return
        cols    = list(tree['columns'])
        col_idx = int(col_id.lstrip('#')) - 1
        if col_idx < 0 or col_idx >= len(cols):
            return
        col_name = cols[col_idx]
        if col_name == 'Group':
            self._start_group_edit(tree, event)
            return
        if not col_name.endswith('kHz'):
            return
        mid = tree.item(row_id, 'values')[0]
        try:
            freq_khz = float(col_name.split()[0])
        except ValueError:
            return
        raw_path = getattr(self, '_thresh_raw_map', {}).get((mid, freq_khz))
        if raw_path:
            self._reanalyze_from_map(mid, raw_path)
        else:
            messagebox.showinfo("Not found",
                                f"No raw file recorded for {mid} at {freq_khz:.2f} kHz.\n"
                                "Try refreshing the table first.")

    def _on_wave1_dclick(self, event):
        """Double-click on Wave 1 table: Group column → edit; data cell → re-analyze."""
        tree   = self.wave1_tree
        row_id = tree.identify_row(event.y)
        col_id = tree.identify_column(event.x)
        if not row_id:
            return
        cols     = list(tree['columns'])
        col_idx  = int(col_id.lstrip('#')) - 1
        if col_idx < 0 or col_idx >= len(cols):
            return
        col_name = cols[col_idx]
        if col_name == 'Group':
            self._start_group_edit(tree, event)
            return
        mid      = tree.item(row_id, 'values')[0]
        raw_path = getattr(self, '_wave1_raw_map', {}).get(mid)
        if raw_path:
            self._reanalyze_from_map(mid, raw_path)
        else:
            messagebox.showinfo("Not found",
                                f"No raw file recorded for {mid}.\n"
                                "Try refreshing the table first.")

    # ===========================================================
    # Tab 1 – ABR Thresholds
    # ===========================================================

    def refresh_threshold_table(self):
        if not self.mice_data:
            return

        thresh_dict = {}
        all_freqs   = set()
        self._thresh_raw_map = {}   # {(mid, freq_khz): raw_filepath}

        for mid, mdata in self.mice_data.items():
            # Skip DP-only entries — they belong only in the DPOAE tab
            if not mdata['analyzed_files'] and not mdata['raw_abr_files']:
                continue
            thresh_dict[mid] = {}
            for fp in mdata['analyzed_files']:
                p = self._get_analyzed(fp)
                if p['frequency'] is not None:
                    all_freqs.add(p['frequency'])
                    thresh_dict[mid][p['frequency']] = p['threshold']  # may be None
                    raw_path = fp.replace('-analyzed.txt', '')
                    self._thresh_raw_map[(mid, p['frequency'])] = raw_path

        freqs = sorted(all_freqs)
        cols  = ['Mouse', 'Group'] + [f"{f:.2f} kHz" for f in freqs]
        self._configure_tree(self.thresh_tree, cols)

        for mid in sorted(thresh_dict.keys()):
            group = self._get_group(mid)
            row   = [mid, group]
            for f in freqs:
                v = thresh_dict[mid].get(f)
                row.append(f"{v:.0f}" if v is not None else 'NaN')
            self.thresh_tree.insert('', tk.END, values=row)

        self._thresh_plot_data = thresh_dict
        self._refresh_thresh_plot()

    def _refresh_thresh_plot(self):
        """Redraw the ABR audiogram (threshold vs frequency) plot."""
        ax = self._thresh_ax
        ax.clear()

        data = getattr(self, '_thresh_plot_data', {})
        if not data:
            ax.text(0.5, 0.5, 'No data — click Refresh',
                    ha='center', va='center', transform=ax.transAxes, color='gray')
            self._thresh_canvas.draw()
            return

        mode      = self._thresh_plot_mode.get()
        all_freqs = sorted({f for d in data.values() for f in d})
        palette   = ['#2196F3', '#E53935', '#43A047', '#FB8C00',
                     '#8E24AA', '#00ACC1', '#6D4C41', '#FFB300']

        group_data = {}
        for mid, d in data.items():
            grp = self._get_group(mid)
            if not grp.strip() or grp.strip().upper() == 'N/A':  # excluded from plot
                continue
            group_data.setdefault(grp, {})[mid] = d

        # Degenerate-grouping fix: pool into one grand mean when every animal
        # is its own unique group (e.g. no shared parent folder).
        if mode == 'mean':
            has_real_groups = any(len(mice) > 1 for mice in group_data.values())
            if not has_real_groups and len(group_data) > 1:
                group_data = {'All Mice': {mid: d for mid, d in data.items()}}

        group_names = sorted(group_data.keys())
        color_map   = {g: palette[i % len(palette)] for i, g in enumerate(group_names)}

        if mode == 'mean':
            for grp in group_names:
                mice_dicts = list(group_data[grp].values())
                means, sems, valid_freqs = [], [], []
                for f in all_freqs:
                    vals = [d[f] for d in mice_dicts
                            if f in d and d[f] is not None]
                    if not vals:
                        continue
                    m = float(np.nanmean(vals))
                    s = (float(np.nanstd(vals, ddof=1)) / np.sqrt(np.sum(~np.isnan(vals)))
                         if len(vals) > 1 else 0.0)
                    means.append(m); sems.append(s); valid_freqs.append(f)
                if not valid_freqs:
                    continue
                c = color_map[grp]
                n = len(mice_dicts)
                ax.plot(valid_freqs, means, '-o', color=c,
                        label=f"{grp} (n={n})", linewidth=2, markersize=5)
                ax.fill_between(valid_freqs,
                                [m - s for m, s in zip(means, sems)],
                                [m + s for m, s in zip(means, sems)],
                                alpha=0.2, color=c)

        else:  # individual — thin traces per animal + thick group mean
            for grp in group_names:
                c          = color_map[grp]
                mice_dicts = list(group_data[grp].values())
                for mid, d in sorted(group_data[grp].items()):
                    fq  = sorted(f for f in d if d[f] is not None)
                    thr = [d[f] for f in fq]
                    ax.plot(fq, thr, '-', color=c, alpha=0.3, linewidth=1.0)
                means, valid_freqs = [], []
                for f in all_freqs:
                    vals = [d[f] for d in mice_dicts
                            if f in d and d[f] is not None]
                    if not vals:
                        continue
                    means.append(float(np.nanmean(vals)))
                    valid_freqs.append(f)
                if valid_freqs:
                    n = len(mice_dicts)
                    ax.plot(valid_freqs, means, '-o', color=c, linewidth=2.5,
                            markersize=6, label=f"{grp} mean (n={n})", zorder=5)

        # Log2 frequency axis with plain numeric tick labels (4, 8, 16, 32 …)
        ax.set_xscale('log', base=2)
        ax.set_xticks(all_freqs)
        ax.xaxis.set_major_formatter(matplotlib.ticker.ScalarFormatter())
        ax.xaxis.set_minor_formatter(matplotlib.ticker.NullFormatter())
        ax.set_xlim(min(all_freqs) * 0.85, max(all_freqs) * 1.15)

        ax.set_xlabel("Frequency (kHz)", fontsize=9)
        ax.set_ylabel("Threshold (dB SPL)", fontsize=9)
        ax.set_title("ABR Audiogram", fontsize=10)
        ax.grid(True, alpha=0.3, linestyle='--')
        ax.tick_params(labelsize=8)

        handles, labels = ax.get_legend_handles_labels()
        if handles:
            ax.legend(fontsize=8, ncol=2 if len(handles) > 8 else 1,
                      framealpha=0.8)

        self._thresh_fig.tight_layout()
        self._thresh_canvas.draw()

    def _save_thresh_plot(self):
        """Save the ABR threshold (audiogram) plot to file."""
        fp = filedialog.asksaveasfilename(
            defaultextension='.png',
            filetypes=[('PNG image', '*.png'),
                       ('PDF',       '*.pdf'),
                       ('SVG',       '*.svg')],
            title="Save Threshold Plot")
        if fp:
            self._thresh_fig.savefig(fp, dpi=150, bbox_inches='tight')
            messagebox.showinfo("Saved", f"Figure saved:\n{fp}")

    def _save_wave1_plot(self):
        """Save the Wave 1 growth function plot to file."""
        fp = filedialog.asksaveasfilename(
            defaultextension='.png',
            filetypes=[('PNG image', '*.png'),
                       ('PDF',       '*.pdf'),
                       ('SVG',       '*.svg')],
            title="Save Wave 1 Plot")
        if fp:
            self._wave1_fig.savefig(fp, dpi=150, bbox_inches='tight')
            messagebox.showinfo("Saved", f"Figure saved:\n{fp}")

    # ===========================================================
    # Tab 2 – Wave 1 Growth Functions
    # ===========================================================

    def refresh_wave1_table(self):
        if not self.mice_data:
            return
        try:
            target_freq = float(self.wave1_freq_var.get())
        except (ValueError, tk.TclError):
            return

        try:
            wn = int(self._wave_num_var.get())
        except (ValueError, tk.TclError):
            wn = 1
        p_lat_col = f'P{wn} Latency'
        n_lat_col = f'N{wn} Latency'
        p_amp_col = f'P{wn} Amplitude'
        n_amp_col = f'N{wn} Amplitude'

        wave1_dict = {}
        all_levels  = set()
        self._wave1_raw_map = {}   # {mid: raw_filepath} for the current frequency

        for mid, mdata in self.mice_data.items():
            # Skip DP-only entries — they belong only in the DPOAE tab
            if not mdata['analyzed_files'] and not mdata['raw_abr_files']:
                continue
            for fp in mdata['analyzed_files']:
                p = self._get_analyzed(fp)
                if p['frequency'] is None:
                    continue
                if abs(p['frequency'] - target_freq) > 0.1:
                    continue
                if p['data'] is None:
                    continue

                df = p['data']
                wave1_dict[mid] = {}
                self._wave1_raw_map[mid] = fp.replace('-analyzed.txt', '')

                for _, row in df.iterrows():
                    level   = float(row['Level'])
                    p_lat   = row.get(p_lat_col, np.nan)
                    n_lat   = row.get(n_lat_col, np.nan)
                    p_amp   = row.get(p_amp_col, np.nan)
                    n_amp   = row.get(n_amp_col, np.nan)

                    all_levels.add(level)   # track every tested level
                    if (not np.isnan(p_lat) and p_lat > 0 and
                            not np.isnan(n_lat) and n_lat > 0 and
                            not np.isnan(p_amp) and not np.isnan(n_amp)):
                        wave1_dict[mid][level] = round(p_amp - n_amp, 4)
                    else:
                        wave1_dict[mid][level] = np.nan  # sub-threshold

        if not wave1_dict:
            return

        levels = sorted(all_levels, reverse=True)
        cols   = ['Mouse', 'Group'] + [f"{l:.0f} dB" for l in levels]
        self._configure_tree(self.wave1_tree, cols)

        for mid in sorted(wave1_dict.keys()):
            group = self._get_group(mid)
            row   = [mid, group]
            for lv in levels:
                if lv not in wave1_dict[mid]:
                    row.append('NM')           # level not tested for this animal
                elif np.isnan(wave1_dict[mid][lv]):
                    row.append('NaN')          # tested but sub-threshold
                else:
                    row.append(f"{wave1_dict[mid][lv]:.4f}")
            self.wave1_tree.insert('', tk.END, values=row)

        # Update plot
        self._wave1_plot_data = wave1_dict
        self._refresh_wave1_plot()

    def _refresh_wave1_plot(self):
        """Redraw the Wave 1 growth function plot."""
        ax = self._wave1_ax
        ax.clear()

        data = getattr(self, '_wave1_plot_data', {})
        if not data:
            ax.text(0.5, 0.5, 'No data — click Refresh',
                    ha='center', va='center', transform=ax.transAxes, color='gray')
            self._wave1_canvas.draw()
            return

        mode = self._wave1_plot_mode.get()
        all_levels = sorted({lv for d in data.values() for lv in d})

        # One colour per group
        palette = ['#2196F3', '#E53935', '#43A047', '#FB8C00',
                   '#8E24AA', '#00ACC1', '#6D4C41', '#FFB300']
        group_data = {}
        for mid, d in data.items():
            grp = self._get_group(mid)
            if not grp.strip() or grp.strip().upper() == 'N/A':  # excluded from plot
                continue
            group_data.setdefault(grp, {})[mid] = d

        # In mean mode: if every "group" has only 1 mouse (degenerate grouping
        # because every animal has a unique parent folder), pool all mice into
        # a single group so we can compute a meaningful grand mean ± SEM.
        if mode == 'mean':
            has_real_groups = any(len(mice) > 1 for mice in group_data.values())
            if not has_real_groups and len(group_data) > 1:
                group_data = {'All Mice': {mid: d for mid, d in data.items()}}

        group_names = sorted(group_data.keys())
        color_map   = {g: palette[i % len(palette)] for i, g in enumerate(group_names)}

        if mode == 'mean':
            for grp in group_names:
                mice_dicts = list(group_data[grp].values())
                means, sems, valid_lvs = [], [], []
                for lv in all_levels:
                    vals = [d[lv] for d in mice_dicts if lv in d]
                    if not vals:
                        continue
                    m = float(np.nanmean(vals))
                    s = (float(np.nanstd(vals, ddof=1)) / np.sqrt(np.sum(~np.isnan(vals)))
                         if len(vals) > 1 else 0.0)
                    means.append(m); sems.append(s); valid_lvs.append(lv)
                if not valid_lvs:
                    continue
                c = color_map[grp]
                n = len(mice_dicts)
                ax.plot(valid_lvs, means, '-o', color=c,
                        label=f"{grp} (n={n})", linewidth=2, markersize=5)
                ax.fill_between(valid_lvs,
                                [m - s for m, s in zip(means, sems)],
                                [m + s for m, s in zip(means, sems)],
                                alpha=0.2, color=c)

        else:  # individual — thin traces per animal + thick group mean
            for grp in group_names:
                c          = color_map[grp]
                mice_dicts = list(group_data[grp].values())
                for mid, d in sorted(group_data[grp].items()):
                    lvs  = sorted(d.keys())
                    amps = [d[lv] for lv in lvs]
                    ax.plot(lvs, amps, '-', color=c, alpha=0.3, linewidth=1.0)
                means, valid_lvs = [], []
                for lv in all_levels:
                    vals = [d[lv] for d in mice_dicts if lv in d]
                    if not vals:
                        continue
                    means.append(float(np.nanmean(vals)))
                    valid_lvs.append(lv)
                if valid_lvs:
                    n = len(mice_dicts)
                    ax.plot(valid_lvs, means, '-o', color=c, linewidth=2.5,
                            markersize=6, label=f"{grp} mean (n={n})", zorder=5)

        try:
            freq = float(self.wave1_freq_var.get())
            wn   = int(getattr(self, '_wave_num_var', tk.StringVar(value='1')).get())
            ax.set_title(f"Wave {wn} Growth  —  {freq:.2f} kHz", fontsize=10)
        except (ValueError, AttributeError):
            wn = 1
            ax.set_title("Wave Growth Function", fontsize=10)

        ax.set_xlabel("Level (dB SPL)", fontsize=9)
        ax.set_ylabel(f"Wave {wn} Amplitude (\u03bcV)", fontsize=9)
        ax.grid(True, alpha=0.3, linestyle='--')
        ax.tick_params(labelsize=8)

        handles, labels = ax.get_legend_handles_labels()
        if handles:
            ax.legend(fontsize=8, ncol=2 if len(handles) > 8 else 1, framealpha=0.8)

        self._wave1_fig.tight_layout()
        self._wave1_canvas.draw()

    # ===========================================================
    # Tab 3 – Plot Traces
    # ===========================================================

    def plot_traces(self):
        sel = self.mice_listbox.curselection()
        if not sel:
            messagebox.showwarning("No mice selected",
                                   "Please select at least one mouse from the list.")
            return

        selected_mice = [self.mice_listbox.get(i) for i in sel]

        try:
            target_freq = float(self.plot_freq_var.get())
        except (ValueError, tk.TclError):
            messagebox.showwarning("No frequency", "Please select a frequency.")
            return

        lev_sel = self.levels_listbox.curselection()
        if not lev_sel:
            messagebox.showwarning("No levels selected",
                                   "Please select at least one level (Ctrl+click for multiple).")
            return

        target_levels = sorted(
            [float(self.levels_listbox.get(i)) for i in lev_sel],
            reverse=True)

        # Gather waveforms from raw files (pooled and per-group)
        waveforms_by_level = {lv: [] for lv in target_levels}
        # bygroup: {grp: {lv: [waveform arrays]}}
        waveforms_by_grp_lv: dict = {}
        sample_rate_us = 40.0

        for mid in selected_mice:
            grp   = self._get_group(mid) or ''   # blank = unassigned, still pooled for grand mean
            mdata = self.mice_data.get(mid, {})
            for fp in mdata.get('raw_abr_files', []):
                p = self._get_raw(fp)
                if p['frequency'] is None:
                    continue
                if abs(p['frequency'] - target_freq) > 0.1:
                    continue
                if p['waveforms'] is None:
                    continue
                sample_rate_us = p['sample_rate_us']
                for lv in target_levels:
                    if lv in p['levels']:
                        idx  = p['levels'].index(lv)
                        wave = p['waveforms'][:, idx]
                        waveforms_by_level[lv].append(wave)
                        waveforms_by_grp_lv.setdefault(grp, {}).setdefault(lv, []).append(wave)

        # Check for data
        valid_levels = [lv for lv in target_levels if waveforms_by_level[lv]]
        if not valid_levels:
            messagebox.showwarning(
                "No waveform data",
                "No raw waveform data was found for the selected mice/frequency/levels.\n\n"
                "Ensure that raw ABR files (named e.g. ABR-1-1, without '-analyzed') "
                "are present alongside the analyzed files.")
            return

        self.plot_fig.clear()
        mode   = self.plot_mode_var.get()
        colors = [f"C{i}" for i in range(len(valid_levels))]

        palette_grp = ['#2196F3', '#E53935', '#43A047', '#FB8C00',
                       '#8E24AA', '#00ACC1', '#6D4C41', '#FFB300']
        grp_names   = sorted(waveforms_by_grp_lv.keys())
        grp_colors  = {g: palette_grp[i % len(palette_grp)]
                       for i, g in enumerate(grp_names)}

        # ── helper: build stacked axes ────────────────────────────────
        def _stacked_axes(n):
            axs = self.plot_fig.subplots(n, 1, sharex=True)
            return [axs] if n == 1 else list(axs)

        if mode == 'stacked':
            # Grand mean ± SEM, one subplot per level
            axes = _stacked_axes(len(valid_levels))
            for ax_i, lv in enumerate(valid_levels):
                waves   = waveforms_by_level[lv]
                ax      = axes[ax_i]
                ml      = min(len(w) for w in waves)
                arr     = np.array([w[:ml] for w in waves])
                t       = np.arange(ml) * sample_rate_us / 1000.0
                mean_w  = arr.mean(axis=0)
                sem_w   = arr.std(axis=0) / np.sqrt(len(waves))

                ax.fill_between(t, mean_w - sem_w, mean_w + sem_w,
                                alpha=0.25, color=colors[ax_i])
                ax.plot(t, mean_w, color=colors[ax_i], linewidth=1.4)
                ax.axhline(0, color='#888', linewidth=0.5, linestyle='--')
                ax.set_ylabel(f"{lv:.0f} dB\n(\u03bcV)", fontsize=8)
                ax.set_xlim(0, t[-1])
                ax.tick_params(labelsize=7)
                ax.text(0.98, 0.88, f"n={len(waves)}",
                        transform=ax.transAxes, fontsize=7,
                        ha='right', va='top', color='gray')
            axes[-1].set_xlabel("Time (ms)", fontsize=9)

        elif mode == 'overlay':
            # Grand mean ± SEM, all levels on one plot
            ax = self.plot_fig.add_subplot(111)
            for i, lv in enumerate(valid_levels):
                waves  = waveforms_by_level[lv]
                ml     = min(len(w) for w in waves)
                arr    = np.array([w[:ml] for w in waves])
                t      = np.arange(ml) * sample_rate_us / 1000.0
                mean_w = arr.mean(axis=0)
                sem_w  = arr.std(axis=0) / np.sqrt(len(waves))

                ax.fill_between(t, mean_w - sem_w, mean_w + sem_w,
                                alpha=0.2, color=colors[i])
                ax.plot(t, mean_w, color=colors[i], linewidth=1.4,
                        label=f"{lv:.0f} dB (n={len(waves)})")

            ax.axhline(0, color='#888', linewidth=0.5, linestyle='--')
            ax.set_xlabel("Time (ms)", fontsize=9)
            ax.set_ylabel("Amplitude (\u03bcV)", fontsize=9)
            ax.set_xlim(0, t[-1])
            ax.legend(fontsize=8, loc='upper right')

        elif mode == 'bygroup':
            # Stacked by level; each group gets its own coloured mean ± SEM line
            axes       = _stacked_axes(len(valid_levels))
            legend_shown: set = set()
            for ax_i, lv in enumerate(valid_levels):
                ax = axes[ax_i]
                ax.axhline(0, color='#888', linewidth=0.5, linestyle='--')
                for grp in grp_names:
                    waves = waveforms_by_grp_lv.get(grp, {}).get(lv, [])
                    if not waves:
                        continue
                    c      = grp_colors[grp]
                    ml     = min(len(w) for w in waves)
                    arr    = np.array([w[:ml] for w in waves])
                    t      = np.arange(ml) * sample_rate_us / 1000.0
                    mean_w = arr.mean(axis=0)
                    sem_w  = arr.std(axis=0) / np.sqrt(len(waves))
                    ax.fill_between(t, mean_w - sem_w, mean_w + sem_w,
                                    alpha=0.25, color=c)
                    lbl = (f"{grp} (n={len(waves)})"
                           if grp not in legend_shown else '_nolegend_')
                    ax.plot(t, mean_w, color=c, linewidth=1.8, label=lbl)
                    legend_shown.add(grp)
                ax.set_ylabel(f"{lv:.0f} dB\n(\u03bcV)", fontsize=8)
                ax.set_xlim(0, t[-1])
                ax.tick_params(labelsize=7)
            axes[-1].set_xlabel("Time (ms)", fontsize=9)
            hdls, lbls = axes[0].get_legend_handles_labels()
            if hdls:
                axes[0].legend(fontsize=7, loc='upper right', framealpha=0.8)

        else:  # individuals — grey individual traces + coloured grand mean ± SEM
            axes = _stacked_axes(len(valid_levels))
            for ax_i, lv in enumerate(valid_levels):
                waves  = waveforms_by_level[lv]
                ax     = axes[ax_i]
                ml     = min(len(w) for w in waves)
                arr    = np.array([w[:ml] for w in waves])
                t      = np.arange(ml) * sample_rate_us / 1000.0
                mean_w = arr.mean(axis=0)
                sem_w  = arr.std(axis=0) / np.sqrt(len(waves))

                # Individual traces in grey behind the mean
                for w in waves:
                    ax.plot(t, w[:ml], color='#999999',
                            alpha=0.35, linewidth=0.7, zorder=1)

                # Grand mean + SEM on top
                ax.fill_between(t, mean_w - sem_w, mean_w + sem_w,
                                alpha=0.30, color=colors[ax_i], zorder=2)
                ax.plot(t, mean_w, color=colors[ax_i],
                        linewidth=1.8, zorder=3)
                ax.axhline(0, color='#888', linewidth=0.5,
                           linestyle='--', zorder=0)
                ax.set_ylabel(f"{lv:.0f} dB\n(\u03bcV)", fontsize=8)
                ax.set_xlim(0, t[-1])
                ax.tick_params(labelsize=7)
                ax.text(0.98, 0.88, f"n={len(waves)}",
                        transform=ax.transAxes, fontsize=7,
                        ha='right', va='top', color='gray')
            axes[-1].set_xlabel("Time (ms)", fontsize=9)

        self.plot_fig.suptitle(
            f"ABR Traces  —  {target_freq:.1f} kHz   "
            f"[{', '.join(selected_mice)}]",
            fontsize=10)
        self.plot_fig.tight_layout()
        self.plot_canvas.draw()

    def save_plot_figure(self):
        fp = filedialog.asksaveasfilename(
            defaultextension='.png',
            filetypes=[('PNG image', '*.png'),
                       ('PDF',       '*.pdf'),
                       ('SVG',       '*.svg')],
            title="Save Figure")
        if fp:
            self.plot_fig.savefig(fp, dpi=150, bbox_inches='tight')
            messagebox.showinfo("Saved", f"Figure saved:\n{fp}")

    # ===========================================================
    # Tab 4 – DPOAE Thresholds
    # ===========================================================

    def refresh_dpoae_table(self):
        if not self.mice_data:
            return
        try:
            criterion = float(self.dpoae_criterion.get())
        except (ValueError, tk.TclError):
            criterion = 0.0

        thresh_dict = {}
        all_freqs   = set()

        for mid, mdata in self.mice_data.items():
            thresh_dict[mid] = {}
            for fp in mdata['dp_files']:
                dp_df = self._get_dp(fp)
                if dp_df.empty:
                    continue
                for f2 in dp_df['f2_Hz'].unique():
                    if np.isnan(f2):
                        continue
                    f2_khz = round(f2 / 1000.0, 3)
                    t      = dpoae_threshold(dp_df, f2, criterion)
                    thresh_dict[mid][f2_khz] = t if not np.isnan(t) else np.nan
                    all_freqs.add(f2_khz)

        if not all_freqs:
            return

        freqs = sorted(all_freqs)
        cols  = ['Mouse', 'Group'] + [f"{f:.1f} kHz" for f in freqs]
        self._configure_tree(self.dpoae_tree, cols)

        for mid in sorted(thresh_dict.keys()):
            group = self._get_group(mid)
            row   = [mid, group]
            for f in freqs:
                v = thresh_dict[mid].get(f, 'NM')
                if v == 'NM':
                    row.append('NM')
                elif isinstance(v, float) and np.isnan(v):
                    row.append('NaN')          # no level reached criterion
                else:
                    row.append(f"{v:.0f}")
            self.dpoae_tree.insert('', tk.END, values=row)

        # Store numeric-only data for plotting (skip NaN / NM)
        plot_data = {}
        for mid, fd in thresh_dict.items():
            d = {f: v for f, v in fd.items()
                 if isinstance(v, float) and not np.isnan(v)}
            if d:
                plot_data[mid] = d
        self._dpoae_plot_data = plot_data
        self._refresh_dpoae_plot()

    def _refresh_dpoae_plot(self):
        """Redraw the DPOAE audiogram (threshold vs frequency)."""
        ax = self._dpoae_ax
        ax.clear()

        data = getattr(self, '_dpoae_plot_data', {})
        if not data:
            ax.text(0.5, 0.5, 'No data — click Calculate',
                    ha='center', va='center', transform=ax.transAxes, color='gray')
            self._dpoae_canvas.draw()
            return

        mode      = self._dpoae_plot_mode.get()
        all_freqs = sorted({f for d in data.values() for f in d})
        palette   = ['#2196F3', '#E53935', '#43A047', '#FB8C00',
                     '#8E24AA', '#00ACC1', '#6D4C41', '#FFB300']

        group_data = {}
        for mid, d in data.items():
            grp = self._get_group(mid)
            if not grp.strip() or grp.strip().upper() == 'N/A':  # excluded from plot
                continue
            group_data.setdefault(grp, {})[mid] = d

        if mode == 'mean':
            has_real = any(len(m) > 1 for m in group_data.values())
            if not has_real and len(group_data) > 1:
                group_data = {'All Mice': {m: d for m, d in data.items()}}

        group_names = sorted(group_data.keys())
        color_map   = {g: palette[i % len(palette)] for i, g in enumerate(group_names)}

        if mode == 'mean':
            for grp in group_names:
                mice_dicts = list(group_data[grp].values())
                means, sems, valid_freqs = [], [], []
                for f in all_freqs:
                    vals = [d[f] for d in mice_dicts if f in d]
                    if not vals:
                        continue
                    m = float(np.nanmean(vals))
                    s = (float(np.nanstd(vals, ddof=1)) / np.sqrt(np.sum(~np.isnan(vals)))
                         if len(vals) > 1 else 0.0)
                    means.append(m); sems.append(s); valid_freqs.append(f)
                if not valid_freqs:
                    continue
                c = color_map[grp]
                n = len(mice_dicts)
                ax.plot(valid_freqs, means, '-o', color=c,
                        label=f"{grp} (n={n})", linewidth=2, markersize=5)
                ax.fill_between(valid_freqs,
                                [m - s for m, s in zip(means, sems)],
                                [m + s for m, s in zip(means, sems)],
                                alpha=0.2, color=c)

        else:  # individual
            for grp in group_names:
                c          = color_map[grp]
                mice_dicts = list(group_data[grp].values())
                for mid, d in sorted(group_data[grp].items()):
                    fq  = sorted(d.keys())
                    thr = [d[f] for f in fq]
                    ax.plot(fq, thr, '-', color=c, alpha=0.3, linewidth=1.0)
                means, valid_freqs = [], []
                for f in all_freqs:
                    vals = [d[f] for d in mice_dicts if f in d]
                    if not vals:
                        continue
                    means.append(float(np.nanmean(vals)))
                    valid_freqs.append(f)
                if valid_freqs:
                    n = len(mice_dicts)
                    ax.plot(valid_freqs, means, '-o', color=c, linewidth=2.5,
                            markersize=6, label=f"{grp} mean (n={n})", zorder=5)

        # Log2 frequency axis
        ax.set_xscale('log', base=2)
        ax.set_xticks(all_freqs)
        ax.xaxis.set_major_formatter(matplotlib.ticker.ScalarFormatter())
        ax.xaxis.set_minor_formatter(matplotlib.ticker.NullFormatter())
        ax.set_xlim(min(all_freqs) * 0.85, max(all_freqs) * 1.15)

        ax.set_xlabel("Frequency (kHz)", fontsize=9)
        ax.set_ylabel("Threshold (dB SPL)", fontsize=9)
        ax.set_title("DPOAE Audiogram", fontsize=10)
        ax.grid(True, alpha=0.3, linestyle='--')
        ax.tick_params(labelsize=8)

        handles, labels = ax.get_legend_handles_labels()
        if handles:
            ax.legend(fontsize=8, ncol=2 if len(handles) > 8 else 1, framealpha=0.8)

        self._dpoae_fig.tight_layout()
        self._dpoae_canvas.draw()

    def _save_dpoae_plot(self):
        fp = filedialog.asksaveasfilename(
            defaultextension='.png',
            filetypes=[('PNG image', '*.png'), ('PDF', '*.pdf'), ('SVG', '*.svg')],
            title="Save DPOAE Plot")
        if fp:
            self._dpoae_fig.savefig(fp, dpi=150, bbox_inches='tight')
            messagebox.showinfo("Saved", f"Figure saved:\n{fp}")

    # ===========================================================
    # Excel Export
    # ===========================================================

    def export_thresholds_excel(self):
        if not self.mice_data:
            messagebox.showwarning("No data", "Load a data folder first.")
            return
        fp = filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            filetypes=[('Excel workbook', '*.xlsx')],
            title="Save ABR Thresholds")
        if not fp:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "ABR Thresholds"
            self.refresh_threshold_table()
            self._write_df_to_sheet(ws,
                                    self._tree_to_df(self.thresh_tree),
                                    "ABR Thresholds (dB SPL)")
            wb.save(fp)
            messagebox.showinfo("Saved", f"Saved:\n{fp}")
        except Exception as exc:
            messagebox.showerror("Export error", str(exc))

    def export_wave1_excel(self):
        """Export one sheet per frequency, plus the threshold sheet."""
        if not self.mice_data:
            messagebox.showwarning("No data", "Load a data folder first.")
            return
        fp = filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            filetypes=[('Excel workbook', '*.xlsx')],
            title="Save Wave 1 Growth Functions")
        if not fp:
            return
        try:
            wb = openpyxl.Workbook()

            # Sheet 1: thresholds
            ws_thresh = wb.active
            ws_thresh.title = "ABR Thresholds"
            self.refresh_threshold_table()
            self._write_df_to_sheet(ws_thresh,
                                    self._tree_to_df(self.thresh_tree),
                                    "ABR Thresholds (dB SPL)")

            # One sheet per frequency
            all_freqs = set()
            for mdata in self.mice_data.values():
                for afp in mdata['analyzed_files']:
                    p = self._get_analyzed(afp)
                    if p['frequency'] is not None:
                        all_freqs.add(p['frequency'])

            for freq in sorted(all_freqs):
                self.wave1_freq_var.set(f"{freq:.2f}")
                self.refresh_wave1_table()
                df = self._tree_to_df(self.wave1_tree)
                if df.empty:
                    continue
                sheet_name = f"Wave1_{freq:.1f}kHz"[:31]
                ws = wb.create_sheet(title=sheet_name)
                self._write_df_to_sheet(
                    ws, df,
                    f"Wave 1 Amplitude (μV)  —  {freq:.2f} kHz")

            wb.save(fp)
            messagebox.showinfo("Saved", f"Saved:\n{fp}")
        except Exception as exc:
            messagebox.showerror("Export error", str(exc))

    def export_dpoae_excel(self):
        if not self.mice_data:
            messagebox.showwarning("No data", "Load a data folder first.")
            return
        fp = filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            filetypes=[('Excel workbook', '*.xlsx')],
            title="Save DPOAE Thresholds")
        if not fp:
            return
        try:
            wb  = openpyxl.Workbook()
            ws  = wb.active
            ws.title = "DPOAE Thresholds"
            self.refresh_dpoae_table()
            crit = self.dpoae_criterion.get()
            self._write_df_to_sheet(
                ws,
                self._tree_to_df(self.dpoae_tree),
                f"DPOAE Thresholds (dB SPL)  —  criterion: {crit:.0f} dB SPL  |  NaN = no response")
            wb.save(fp)
            messagebox.showinfo("Saved", f"Saved:\n{fp}")
        except Exception as exc:
            messagebox.showerror("Export error", str(exc))

    # ===========================================================
    # Shared helpers
    # ===========================================================

    def _configure_tree(self, tree, cols):
        tree['columns'] = cols
        tree.delete(*tree.get_children())
        for col in cols:
            tree.heading(col, text=col)
            w = 110 if col in ('Mouse', 'Group') else 78
            tree.column(col, width=w, anchor='center', minwidth=60)

    def _tree_to_df(self, tree):
        cols = tree['columns']
        rows = [tree.item(item)['values'] for item in tree.get_children()]
        return pd.DataFrame(rows, columns=cols)

    def _write_df_to_sheet(self, ws, df, title):
        hdr_font  = Font(bold=True, color='FFFFFF')
        hdr_fill  = PatternFill('solid', fgColor='1E4D7B')
        title_fnt = Font(bold=True, size=12)
        center    = Alignment(horizontal='center')

        ws.cell(row=1, column=1, value=title).font = title_fnt

        for ci, cname in enumerate(df.columns, start=1):
            cell = ws.cell(row=2, column=ci, value=cname)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = center

        for ri, row in df.iterrows():
            for ci, val in enumerate(row, start=1):
                cell = ws.cell(row=ri + 3, column=ci, value=val)
                cell.alignment = center

        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max(10, max_len + 3)

    def _show_about(self):
        messagebox.showinfo(
            "About",
            "CoFAST: COchlear Function Analysis SuiTe\n\n"
            "Developed by the Functional Testing Core at\n"
            "Eaton-Peabody Laboratories (EPL).\n"
            "Integrates with the EPL CFTS acquisition system.\n\n"
            "Features:\n"
            "  • ABR threshold extraction (from *-analyzed.txt files)\n"
            "  • Wave 1 (P1–N1) growth function extraction\n"
            "  • Mean ± SEM waveform plots (from raw ABR files)\n"
            "  • DPOAE threshold extraction (DP-* files)\n\n"
            "Dependencies: numpy, pandas, openpyxl, matplotlib")


# ============================================================
# Peak Analysis Window
# ============================================================

class PeakAnalysisWindow(tk.Toplevel):
    """
    Interactive waveform viewer for manually reviewing and labelling
    ABR peaks on raw (un-analysed) data files.

    One window is shared for all files in the queue; it steps through
    them one at a time (one file = one frequency for one mouse).

    Controls
    --------
    Keyboard:
        T          – set threshold at the current (highlighted) level
        Up / Down  – move current level up / down one step
        1-5        – select peak P1–P5 for placement
        Shift+1-5  – select valley N1–N5 for placement
        Esc        – deselect peak/valley (back to navigation-only mode)
        S / Return – save current file and advance to next

    Mouse (on waveform canvas):
        Left-click – selects the current level (nearest waveform to click)
    """

    # Colours for wave numbers 1-5
    PT_COLORS = {1: '#e74c3c', 2: '#e67e22', 3: '#27ae60',
                 4: '#16a085', 3: '#27ae60', 4: '#16a085', 5: '#2980b9'}

    def __init__(self, parent, file_queue, on_complete=None):
        """
        file_queue  : list of (mouse_id, raw_abr_filepath)
        on_complete : optional callable; invoked after all files are done
        """
        super().__init__(parent)
        self.file_queue    = list(file_queue)
        self.queue_idx     = 0
        self.on_complete   = on_complete

        # ── per-file state ────────────────────────────────────────
        self.mouse_id      = None
        self.raw_path      = None
        self.frequency_khz = None
        self.fs_hz         = None
        self.time_ms       = None
        self.waveform_mat  = None    # (n_samples, n_levels) highest→lowest
        self.diff_mat      = None    # polarity-difference channel (for ABRpresto)
        self.level_list    = []      # floats, highest→lowest
        self.peak_idx      = []      # list[dict {1-5:int}]
        self.valley_idx    = []      # list[dict {1-5:int}]
        self.threshold_db  = None
        self.cur_lv        = 0       # index into level_list
        self._undo_stack   = []      # list of (peak_idx, valley_idx, threshold_db) snapshots
        self.filter_zpk    = "No filtering"

        # ── UI state ──────────────────────────────────────────────
        self.scale_var      = tk.DoubleVar(value=7.0)
        self.norm_var       = tk.BooleanVar(value=False)
        self.point_mode     = None   # None | ('P',1-5) | ('N',1-5)

        self.title("ABR Peak Analysis")
        self.resizable(True, True)
        self.geometry("1300x860")
        self._build_ui()
        self._load_current()

    # ----------------------------------------------------------
    # UI construction
    # ----------------------------------------------------------

    def _build_ui(self):
        self.configure(bg='#f0f0f0')

        # ── progress bar at top ───────────────────────────────────
        self._title_var = tk.StringVar()
        tk.Label(self, textvariable=self._title_var,
                 font=('Arial', 10, 'bold'), bg='#1e4d7b', fg='white',
                 pady=5).pack(fill=tk.X)

        # ── main area: canvas left, controls right ────────────────
        main = tk.Frame(self, bg='#f0f0f0')
        main.pack(fill=tk.BOTH, expand=True)

        # Canvas
        fig_frame = tk.Frame(main, bg='#f0f0f0')
        fig_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self._fig = Figure(figsize=(8, 6), dpi=96)
        self._ax  = self._fig.add_subplot(111)
        self._canvas = FigureCanvasTkAgg(self._fig, master=fig_frame)
        self._canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        self._canvas.mpl_connect('button_press_event', self._on_canvas_click)

        tb_frame = tk.Frame(fig_frame)
        tb_frame.pack(fill=tk.X)
        NavigationToolbar2Tk(self._canvas, tb_frame)

        # Right-side controls
        ctrl = tk.Frame(main, width=190, bg='#f0f0f0',
                        relief=tk.RIDGE, bd=1)
        ctrl.pack(side=tk.RIGHT, fill=tk.Y, padx=4, pady=4)
        ctrl.pack_propagate(False)

        # Scale
        tk.Label(ctrl, text="Waveform Scale", font=('Arial', 9, 'bold'),
                 bg='#f0f0f0').pack(pady=(10, 0))
        self._scale_lbl = tk.Label(ctrl, text="7", bg='#f0f0f0',
                                   font=('Arial', 9))
        self._scale_lbl.pack()
        sc = ttk.Scale(ctrl, from_=1, to=15, orient='horizontal',
                       variable=self.scale_var, command=self._on_scale)
        sc.pack(fill=tk.X, padx=8)
        tk.Checkbutton(ctrl, text="Normalize traces",
                       variable=self.norm_var, command=self._on_normalize_toggle,
                       bg='#f0f0f0', activebackground='#f0f0f0',
                       font=('Arial', 8)).pack(anchor='w', padx=10, pady=(2, 0))

        ttk.Separator(ctrl, orient='horizontal').pack(fill=tk.X,
                                                       padx=6, pady=8)

        # Threshold
        tk.Label(ctrl, text="Threshold (dB SPL)",
                 font=('Arial', 9, 'bold'), bg='#f0f0f0').pack()
        self._thr_var = tk.StringVar(value="—")
        tk.Label(ctrl, textvariable=self._thr_var,
                 font=('Arial', 14, 'bold'), fg='#2980b9',
                 bg='#f0f0f0').pack()
        _mk_btn(ctrl, text="Set Threshold Here  [T]",
                command=self._set_threshold,
                bg='#2980b9', pady=3).pack(fill=tk.X, padx=8, pady=4)

        ttk.Separator(ctrl, orient='horizontal').pack(fill=tk.X,
                                                       padx=6, pady=6)

        # Peak / valley placement selector
        tk.Label(ctrl, text="Place Peak / Valley",
                 font=('Arial', 9, 'bold'), bg='#f0f0f0').pack()
        tk.Label(ctrl, text="(select then use ←/→ to position)",
                 font=('Arial', 8), fg='gray', bg='#f0f0f0').pack()

        self._mode_var = tk.StringVar(value='none')
        btn_grid = tk.Frame(ctrl, bg='#f0f0f0')
        btn_grid.pack(pady=4, padx=4)

        for col, label in enumerate(['P1','P2','P3','P4','P5',
                                      'N1','N2','N3','N4','N5']):
            r, c = divmod(col, 5)
            ttk.Radiobutton(btn_grid, text=label,
                            variable=self._mode_var, value=label,
                            command=self._on_mode_change
                            ).grid(row=r, column=c, sticky='w')

        tk.Button(ctrl, text="✖  None  [Esc]",
                  command=lambda: (self._mode_var.set('none'),
                                   self._on_mode_change()),
                  relief=tk.FLAT, bg='#f0f0f0').pack()

        ttk.Separator(ctrl, orient='horizontal').pack(fill=tk.X,
                                                       padx=6, pady=6)

        _mk_btn(ctrl, text="🔍  Auto-detect All Peaks",
                command=self._auto_detect_all,
                bg='#8e44ad', pady=3).pack(fill=tk.X, padx=8, pady=2)


        ttk.Separator(ctrl, orient='horizontal').pack(fill=tk.X,
                                                       padx=6, pady=6)

        # Keyboard hint
        hint = ("Waveforms\n"
                "  ↑ / ↓       change level\n"
                "  P           invert polarity\n"
                "  N           toggle normalize\n"
                "  + / −       scale up / down\n"
                "\nPeaks\n"
                "  1–5         select P1–P5\n"
                "  ⇧+1–5       select N1–N5\n"
                "  ← / →       snap to next extremum\n"
                "  ⇧+← / →     fine-adjust ±1 sample\n"
                "  I           auto-detect all peaks\n"
                "  U           propagate to lower levels\n"
                "\nThreshold\n"
                "  T           auto-estimate (correlation)\n"
                "  ↵           set to current level\n"
                "\nFile\n"
                "  S           save & next\n"
                "  X           clear & re-detect\n"
                "  R           restore last save\n"
                "  Ctrl+Z      undo\n"
                "  Esc         deselect peak")
        tk.Label(ctrl, text=hint, justify=tk.LEFT,
                 font=('Courier', 8), bg='#f0f0f0',
                 fg='#444').pack(padx=8, pady=(0, 4), anchor='w')

        # ── bottom navigation ─────────────────────────────────────
        nav = tk.Frame(self, bg='#e0e0e0', pady=6)
        nav.pack(fill=tk.X, side=tk.BOTTOM)
        _mk_btn(nav, text="Skip (no save)",
                command=self._skip_current,
                bg='#95a5a6', padx=10).pack(side=tk.LEFT, padx=8)
        _mk_btn(nav, text="Save & Next  →",
                command=self._save_and_next,
                bg='#27ae60', padx=16, pady=4,
                font=('Arial', 10, 'bold')).pack(side=tk.RIGHT, padx=8)

        # Keyboard bindings
        self.bind('<KeyPress>', self._on_key)
        self.focus_set()

    # ----------------------------------------------------------
    # Load / parse one raw ABR file
    # ----------------------------------------------------------

    def _load_current(self):
        if self.queue_idx >= len(self.file_queue):
            self._finish()
            return

        self.mouse_id, self.raw_path = self.file_queue[self.queue_idx]
        total = len(self.file_queue)
        self._title_var.set(
            f"Peak Analysis  —  {self.mouse_id}  |  "
            f"File {self.queue_idx + 1} of {total}: "
            f"{os.path.basename(self.raw_path)}")

        raw = parse_raw_abr_file(self.raw_path)

        if raw['waveforms'] is None or not raw['levels']:
            messagebox.showwarning("Parse error",
                f"Could not read waveform data from:\n{self.raw_path}\n\n"
                "Skipping to next file.")
            self.queue_idx += 1
            self._load_current()
            return

        self.frequency_khz = raw['frequency'] or 0.0
        self.fs_hz         = 1e6 / raw['sample_rate_us']      # µs/sample → Hz
        self.time_ms       = (np.arange(raw['waveforms'].shape[0])
                              * raw['sample_rate_us'] / 1000.0)

        # Sort columns highest → lowest level
        sorted_pairs = sorted(zip(raw['levels'],
                                  range(len(raw['levels']))),
                              reverse=True)
        self.level_list  = [p[0] for p in sorted_pairs]
        col_order        = [p[1] for p in sorted_pairs]
        raw_mat          = raw['waveforms'][:, col_order]

        # Filter every column (sum channel)
        filtered_cols = []
        zpk_str = "No filtering"
        for c in range(raw_mat.shape[1]):
            filt, zpk_str = _butter_bandpass(raw_mat[:, c], self.fs_hz)
            filtered_cols.append(filt)
        self.waveform_mat = np.column_stack(filtered_cols)
        self.filter_zpk   = zpk_str

        # Filter difference channel (polarity subaverages) if present
        raw_diff = raw.get('diff_waveforms')
        if raw_diff is not None and raw_diff.shape == raw['waveforms'].shape:
            diff_sorted = raw_diff[:, col_order]
            diff_filt_cols = []
            for c in range(diff_sorted.shape[1]):
                filt_d, _ = _butter_bandpass(diff_sorted[:, c], self.fs_hz)
                diff_filt_cols.append(filt_d)
            self.diff_mat = np.column_stack(diff_filt_cols)
        else:
            self.diff_mat = None

        # Start with no peaks — user can auto-detect or place manually
        n_lv = len(self.level_list)
        self.peak_idx   = [{} for _ in range(n_lv)]
        self.valley_idx = [{} for _ in range(n_lv)]

        # No threshold set on load — all levels render as supra until user sets one
        self.threshold_db = None
        self.cur_lv       = 0
        self._thr_var.set("—")
        self._mode_var.set('none')

        self._draw()

    # ----------------------------------------------------------
    # Drawing
    # ----------------------------------------------------------

    def _draw(self):
        self._ax.clear()
        scale    = self.scale_var.get()
        t        = self.time_ms
        thr      = self.threshold_db
        normalize = self.norm_var.get()

        for i, lv in enumerate(self.level_list):
            sig = self.waveform_mat[:, i]
            if normalize:
                peak_amp = np.max(np.abs(sig))
                sig = sig / peak_amp if peak_amp > 0 else sig
            y_disp  = sig * scale + lv
            is_cur  = (i == self.cur_lv)
            is_sup  = (thr is None or lv >= thr)

            if is_cur:
                color, lw, ls, alpha = 'black', 2.0, '-', 1.0
            elif is_sup:
                color, lw, ls, alpha = '#444444', 1.0, '-', 0.9
            else:
                color, lw, ls, alpha = '#aaaaaa', 0.8, '--', 0.6

            self._ax.plot(t, y_disp, color=color, linewidth=lw,
                          linestyle=ls, alpha=alpha)

            # Level label on the right
            self._ax.text(t[-1] + 0.05, lv, f"{lv:.0f}",
                          va='center', ha='left', fontsize=7,
                          color='black' if is_cur else color,
                          fontweight='bold' if is_cur else 'normal')

            # Peak / valley markers  (use `sig` which may be normalised)
            pt_alpha = 1.0 if is_sup else 0.25
            for n in range(1, 6):
                c = self.PT_COLORS.get(n, 'gray')
                if n in self.peak_idx[i]:
                    xi = self.peak_idx[i][n]
                    self._ax.plot(t[xi],
                                  sig[xi] * scale + lv,
                                  'o', color=c, markersize=5,
                                  alpha=pt_alpha, zorder=5)
                if n in self.valley_idx[i]:
                    xi = self.valley_idx[i][n]
                    self._ax.plot(t[xi],
                                  sig[xi] * scale + lv,
                                  'v', color=c, markersize=5,
                                  alpha=pt_alpha, zorder=5)

        # Threshold marker
        if thr is not None:
            self._ax.axhline(thr, color='#2980b9', linewidth=0.8,
                             linestyle=':', alpha=0.7)
            self._ax.text(-0.4, thr, f'▶ {thr:.0f}',
                          va='center', ha='right', fontsize=7,
                          color='#2980b9', fontweight='bold')

        # Highlight current-mode label
        mode = self._mode_var.get()
        mode_txt = f"Placing: {mode}" if mode != 'none' else "Navigate mode"
        self._ax.set_title(
            f"{self.mouse_id}  —  {self.frequency_khz:.2f} kHz   "
            f"[{mode_txt}]",
            fontsize=10)
        self._ax.set_xlabel("Time (ms)", fontsize=9)
        self._ax.set_ylabel("dB SPL  (waveforms offset by level)", fontsize=9)
        t_end = float(t[-1])
        self._ax.set_xlim(-0.5, t_end + 0.3)
        self._ax.set_yticks(self.level_list)
        self._ax.set_yticklabels([f"{l:.0f}" for l in self.level_list],
                                  fontsize=7)
        self._ax.yaxis.set_tick_params(labelsize=7)
        self._fig.tight_layout()
        self._canvas.draw_idle()

    # ----------------------------------------------------------
    # Event handlers
    # ----------------------------------------------------------

    def _on_canvas_click(self, event):
        if event.inaxes is None or event.ydata is None:
            return

        # Click only selects the current level — peaks are moved with arrow keys
        lv_arr = np.array(self.level_list)
        self.cur_lv = int(np.argmin(np.abs(lv_arr - event.ydata)))
        self._draw()

    def _on_key(self, event):
        key  = event.keysym
        mode = self._mode_var.get()

        if key == 'Up':
            self.cur_lv = max(0, self.cur_lv - 1)
            self._draw()
        elif key == 'Down':
            self.cur_lv = min(len(self.level_list) - 1, self.cur_lv + 1)
            self._draw()
        elif key in ('Left', 'Right') and mode != 'none':
            self._push_undo()
            n      = int(mode[1])
            n_samp = len(self.time_ms)
            is_p   = mode[0] == 'P'
            store  = self.peak_idx[self.cur_lv] if is_p else self.valley_idx[self.cur_lv]
            cur    = store.get(n, 0)
            shift  = bool(event.state & 0x1)   # Shift held = fine (1 sample)
            if shift:
                # Fine-adjust: move exactly 1 sample
                step  = -1 if key == 'Left' else 1
                store[n] = int(np.clip(cur + step, 0, n_samp - 1))
            else:
                # Coarse: snap to next NZC extremum in travel direction
                sig    = self.waveform_mat[:, self.cur_lv]
                nzcs_p = _nzc_peaks(sig, self.time_ms, min_latency_ms=0.0)
                nzcs_n = _nzc_peaks(-sig, self.time_ms, min_latency_ms=0.0)
                cands  = sorted(set(nzcs_p if is_p else nzcs_n))
                if key == 'Right':
                    nexts = [c for c in cands if c > cur]
                    store[n] = int(nexts[0]) if nexts else int(np.clip(cur + 1, 0, n_samp - 1))
                else:
                    prevs = [c for c in cands if c < cur]
                    store[n] = int(prevs[-1]) if prevs else int(np.clip(cur - 1, 0, n_samp - 1))
            self._draw()
        elif key == 'p':
            # Polarity flip
            self._push_undo()
            self.waveform_mat = -self.waveform_mat
            # Swap peak ↔ valley indices (positivity becomes negativity)
            self.peak_idx, self.valley_idx = self.valley_idx, self.peak_idx
            self._draw()
        elif key in ('plus', 'equal', 'KP_Add'):
            self.scale_var.set(round(self.scale_var.get() * 1.3, 2))
            self._on_scale()
        elif key in ('minus', 'KP_Subtract', 'underscore'):
            self.scale_var.set(round(self.scale_var.get() / 1.3, 2))
            self._on_scale()
        elif key == 'u' and mode != 'none':
            # Propagate current peak/valley position to all lower levels
            self._push_undo()
            n     = int(mode[1])
            is_p  = mode[0] == 'P'
            src   = self.peak_idx[self.cur_lv] if is_p else self.valley_idx[self.cur_lv]
            seed  = src.get(n)
            if seed is not None:
                t_seed = self.time_ms[seed]
                for ci in range(self.cur_lv + 1, len(self.level_list)):
                    sig   = self.waveform_mat[:, ci]
                    store = self.peak_idx[ci] if is_p else self.valley_idx[ci]
                    store[n] = _snap_to_extremum(sig, self.time_ms,
                                                 t_seed, find_max=is_p,
                                                 half_win=0.7)
            self._draw()
        elif key == 'x':
            # Clear and restart with auto-detection
            self._push_undo()
            self._auto_detect_all()
        elif key == 'r':
            # Restore last-saved analysis from disk (if *-analyzed.txt exists)
            saved_path = self.raw_path + '-analyzed.txt'
            if os.path.isfile(saved_path):
                self._push_undo()
                parsed = parse_analyzed_file(saved_path)
                if parsed['data'] is not None:
                    df = parsed['data']
                    for ci, lv in enumerate(self.level_list):
                        row = df[np.isclose(df['Level'], lv)]
                        if row.empty:
                            continue
                        row = row.iloc[0]
                        for nn in range(1, 6):
                            for col, store in ((f'P{nn} Latency', self.peak_idx[ci]),
                                               (f'N{nn} Latency', self.valley_idx[ci])):
                                lat = row.get(col, np.nan)
                                if not np.isnan(lat) and lat > 0:
                                    # Convert latency (ms) back to sample index
                                    idx = int(np.argmin(np.abs(self.time_ms - lat)))
                                    store[nn] = idx
                    if parsed['threshold'] is not None:
                        self.threshold_db = parsed['threshold']
                        self._thr_var.set(f"{self.threshold_db:.0f}")
                    self._draw()
            else:
                messagebox.showinfo("Restore", "No saved analysis found for this file.")
        elif key == 't':
            self._auto_estimate_threshold()
        elif key == 'Escape':
            self._mode_var.set('none')
            self._on_mode_change()
        elif key == 'z' and event.state & 0x4:  # Ctrl+Z
            self._undo()
        elif key == 'i':
            self._push_undo()
            self._auto_detect_all()
        elif key == 'n':
            self.norm_var.set(not self.norm_var.get())
            self._on_normalize_toggle()
        elif key == 's':
            self._save_and_next()
        elif key == 'Return':
            self._set_threshold()
        elif key in ('1', '2', '3', '4', '5'):
            self._mode_var.set(f"P{key}")
            self._on_mode_change()
        elif key in ('exclam', 'at', 'numbersign', 'dollar', 'percent'):
            mapping = {'exclam': '1', 'at': '2', 'numbersign': '3',
                       'dollar': '4', 'percent': '5'}
            self._mode_var.set(f"N{mapping[key]}")
            self._on_mode_change()

    def _push_undo(self):
        import copy
        self._undo_stack.append((
            copy.deepcopy(self.peak_idx),
            copy.deepcopy(self.valley_idx),
            self.threshold_db,
        ))

    def _undo(self):
        if not self._undo_stack:
            return
        self.peak_idx, self.valley_idx, self.threshold_db = self._undo_stack.pop()
        thr = self.threshold_db
        self._thr_var.set(f"{thr:.0f}" if thr is not None else "—")
        self._draw()

    def _on_scale(self, _val=None):
        v = self.scale_var.get()
        self._scale_lbl.config(text=f"{v:.1f}")
        self._draw()

    def _on_normalize_toggle(self):
        if self.norm_var.get():
            self.scale_var.set(self.scale_var.get() / 2)
        else:
            self.scale_var.set(self.scale_var.get() * 2)
        self._scale_lbl.config(text=f"{self.scale_var.get():.1f}")
        self._draw()

    def _on_mode_change(self):
        self._draw()

    # ----------------------------------------------------------
    # Actions
    # ----------------------------------------------------------

    def _set_threshold(self):
        self._push_undo()
        self.threshold_db = self.level_list[self.cur_lv]
        self._thr_var.set(f"{self.threshold_db:.0f}")
        self._draw()

    # ----------------------------------------------------------
    # Correlation-based threshold estimation (Suthakar & Liberman, 2019)
    # ----------------------------------------------------------

    def _estimate_threshold_correlation(self, criterion=0.35):
        """
        Correlation-based threshold estimation via adjacent-level cross-covariance
        (Suthakar & Liberman, 2019).

        For each adjacent pair of levels, compute the normalised Pearson r
        (cross-covariance at lag = 0) between the two averaged waveforms.
        The x-axis value for each pair is the *lower* of the two levels (dB SPL).

        Two models are fit to the r vs. lower-level function:
          Sigmoid:   y = a + (b−a) / (1 + 10^(d·(c−x)))
                     a = min, b = max, c = midpoint (dB), d = slope
          Power law: y = a·x^b + c   (MATLAB 'power2' fit type)

        Threshold = the level where the better-fitting curve crosses `criterion`.
        The result is snapped to the nearest actually-tested level.

        Decision tree:
          ≤ 2 pairs below criterion  →  response at all levels; return lowest.
          ≤ 2 pairs above criterion  →  no detectable response; return None.
          Otherwise: fit + interpolate/extrapolate.

        Works with any file format — only the averaged waveforms are required.
        """
        from scipy.optimize import curve_fit

        levels = np.array(self.level_list)   # highest → lowest
        if len(levels) < 3:
            return None

        # --- Adjacent-level Pearson r ---
        # levels[i-1] > levels[i]; x = lower level = levels[i]
        r_vals       = []
        lower_levels = []
        for i in range(1, len(levels)):
            r = _compute_corrcoef(self.waveform_mat[:, i],
                                  self.waveform_mat[:, i - 1],
                                  self.time_ms)
            if not np.isnan(r):
                r_vals.append(r)
                lower_levels.append(float(levels[i]))

        if len(r_vals) < 3:
            return None

        sort_idx   = np.argsort(lower_levels)
        lev_sorted = np.array(lower_levels)[sort_idx]
        r_sorted   = np.array(r_vals)[sort_idx]

        # Decision tree
        n_below = int(np.sum(r_sorted <  criterion))
        n_above = int(np.sum(r_sorted >= criterion))
        if n_below <= 2:
            return float(lev_sorted[0])   # response at every tested level
        if n_above <= 2:
            return None                   # no detectable response

        # Model 1 — sigmoid: y = a + (b−a) / (1 + 10^(d·(c−x)))
        def sigmoid(x, a, b, c, d):
            return a + (b - a) / (1.0 + 10.0 ** (d * (c - x)))

        # Model 2 — power law: y = a·x^b + c  (x clamped to ≥1 for safety)
        def power_law(x, a, b, c):
            return a * np.clip(x, 1.0, None) ** b + c

        best_thr  = None
        best_rmse = np.inf
        x_fine    = np.linspace(lev_sorted[0], lev_sorted[-1], 10_000)

        # Sigmoid fit
        try:
            a0, b0 = float(np.min(r_sorted)), float(np.max(r_sorted))
            popt, _ = curve_fit(
                sigmoid, lev_sorted, r_sorted,
                p0=[a0, b0, float(np.median(lev_sorted)), 0.1],
                maxfev=5000,
                bounds=([-1, -1, lev_sorted[0] - 30, 1e-4],
                        [ 1,  2, lev_sorted[-1] + 30, 2.0]))
            rmse   = float(np.sqrt(np.mean((sigmoid(lev_sorted, *popt) - r_sorted) ** 2)))
            y_fine = sigmoid(x_fine, *popt)
            cross  = np.where(np.diff(np.sign(y_fine - criterion)))[0]
            if cross.size and rmse < best_rmse:
                best_rmse = rmse
                best_thr  = float(x_fine[cross[-1]])
        except Exception:
            pass

        # Power-law fit
        try:
            popt, _ = curve_fit(
                power_law, lev_sorted, r_sorted,
                p0=[0.001, 1.5, 0.0], maxfev=5000,
                bounds=([-10, 0.01, -1], [10, 10, 1]))
            rmse   = float(np.sqrt(np.mean((power_law(lev_sorted, *popt) - r_sorted) ** 2)))
            y_fine = power_law(x_fine, *popt)
            cross  = np.where(np.diff(np.sign(y_fine - criterion)))[0]
            if cross.size and rmse < best_rmse:
                best_rmse = rmse
                best_thr  = float(x_fine[cross[-1]])
        except Exception:
            pass

        if best_thr is None:
            return None

        # Snap to nearest tested level
        nearest = int(np.argmin(np.abs(np.array(self.level_list) - best_thr)))
        return float(self.level_list[nearest])

    def _auto_estimate_threshold(self):
        """T key: run correlation-based threshold algorithm; fall back to current level."""
        self._push_undo()
        thr = self._estimate_threshold_correlation()
        if thr is not None:
            self.threshold_db = thr
            self._thr_var.set(f"{thr:.0f}")
            base = self._title_var.get().split('  ✦')[0]
            self._title_var.set(base + f'  ✦ auto-threshold: {thr:.0f} dB')
        else:
            self.threshold_db = self.level_list[self.cur_lv]
            self._thr_var.set(f"{self.threshold_db:.0f}")
            base = self._title_var.get().split('  ✦')[0]
            self._title_var.set(base + '  ✦ threshold set to current level')
        self._draw()

    def _auto_detect_all(self):
        """Re-run auto-detection on every level (seeded from highest)."""
        seed_p = seed_v = None
        for ci in range(len(self.level_list)):
            p, v = _auto_detect_peaks(self.waveform_mat[:, ci],
                                      self.time_ms, seed_p, seed_v)
            self.peak_idx[ci]   = p
            self.valley_idx[ci] = v
            seed_p, seed_v = p, v
        self._draw()

    def _clear_current_peaks(self):
        self._push_undo()
        self.peak_idx[self.cur_lv]   = {}
        self.valley_idx[self.cur_lv] = {}
        self._draw()

    # ----------------------------------------------------------
    # Save / navigation
    # ----------------------------------------------------------

    def _save_and_next(self):
        out_path = self.raw_path + '-analyzed.txt'
        try:
            _write_analyzed_file(
                out_path,
                self.frequency_khz,
                self.threshold_db,
                self.time_ms,
                self.waveform_mat,
                self.level_list,
                self.peak_idx,
                self.valley_idx,
                filter_zpk_str=self.filter_zpk)
        except Exception as exc:
            messagebox.showerror("Save error",
                f"Could not save {os.path.basename(out_path)}:\n{exc}")
            return

        # Add newly created file to the parent app's cache
        if self.on_complete:
            self.on_complete(self.mouse_id, out_path)

        self.queue_idx += 1
        self._load_current()

    def _skip_current(self):
        self.queue_idx += 1
        self._load_current()

    def _finish(self):
        messagebox.showinfo("Peak analysis complete",
            "All files in the queue have been processed.\n"
            "The main window has been refreshed.")
        self.destroy()


# ============================================================
# Entry point
# ============================================================

def main():
    root = tk.Tk()

    # High-DPI awareness on Windows
    try:
        from ctypes import windll
        windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        pass

    app = ABRAnalysisApp(root)
    root.mainloop()


if __name__ == '__main__':
    main()
